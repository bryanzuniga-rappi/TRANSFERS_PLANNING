from __future__ import annotations

from collections import Counter, defaultdict
from contextlib import redirect_stdout
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo
import html
import io
import math
import re
import shutil
import tempfile
import urllib.error
import urllib.request
import zipfile

import openpyxl
import streamlit as st
from reportlab.lib import colors as pdf_colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

import modelo_abasto as engine


APP_NAME = "Transfer Planner"
MAX_UPLOAD_MB = 500
DATA_TRANSFERS_SPREADSHEET_ID = "18kHevkMvf9l4s6ANg3h5KdNyj2yEPGAp5C_t8JwxFVw"

ORIGIN_WAREHOUSES = {
    444: "CITYPARK TURBO",
    831: "CITYPARK CHEDRAUI",
    811: "CEDA TURBO",
    834: "CEDA CHEDRAUI",
    425: "CEDIS LOCAL GDL",
    856: "CEDIS - LOCAL MTY",
    49: "NODO ALTAVISTA",
}

ALEPH_SHEETS = {
    "CATALOGO",
    "NO_DISPONIBLE",
    "STOCK",
    "INSUMOS",
    "GOLDEN_INFALTABLES",
    "TIENDA",
    "STORAGE",
}

MANUAL_BACKEND_SHEETS = {"TIENDAS_CERRADAS"}

REQUIRED_DATABASE_SHEETS = (
    "TIENDAS_CERRADAS",
    "VOLUMETRIA",
    "BLOQUEOS",
    "RUTA_COSTOS",
    "PRIORIDAD",
    "HIGH_VALUE",
    "RACKEADOS",
    "CAP_RECIBO",
    "COPERNICO",
    "CATALOGO",
    "NO_DISPONIBLE",
    "STOCK",
    "INSUMOS",
    "GOLDEN_INFALTABLES",
    "TIENDA",
    "STORAGE",
)

SHEET_DESCRIPTIONS = {
    "TIENDAS_CERRADAS": (
        "Lista permanente de warehouses destino bloqueados. Los requerimientos de "
        "estas tiendas se eliminan automáticamente antes de asignar stock y no "
        "pueden reactivarse desde la interfaz."
    ),
    "VOLUMETRIA": (
        "Volumen en metros cúbicos por unidad de cada SKU. Se utiliza para "
        "calcular el consumo de capacidad de recibo de las tiendas."
    ),
    "BLOQUEOS": (
        "Productos con restricciones regionales de envío, especialmente desde "
        "orígenes de CDMX hacia Guadalajara o Monterrey."
    ),
    "RUTA_COSTOS": (
        "Combinaciones de tienda destino y producto que no cuentan con una ruta "
        "de costos habilitada para la transferencia."
    ),
    "PRIORIDAD": (
        "Prioridad de atención por warehouse destino. El valor 1 representa la "
        "prioridad más alta."
    ),
    "HIGH_VALUE": (
        "Clasificación referencial de productos de alto valor utilizada en los "
        "archivos de salida."
    ),
    "RACKEADOS": (
        "Productos rackeados por warehouse. Para el origen 444, estos productos "
        "se excluyen completamente del stock utilizable."
    ),
    "CAP_RECIBO": (
        "Capacidad máxima de recibo de cada tienda expresada en metros cúbicos."
    ),
    "COPERNICO": (
        "Inventario detallado por ubicación del warehouse 444. Permite identificar "
        "y descontar el stock no utilizable."
    ),
    "CATALOGO": (
        "Catálogo de surtido por tienda y producto con su ADU. Se utiliza para "
        "detectar oportunidades de stockout y completar tareas disponibles mediante "
        "la cobertura AVL opcional."
    ),
    "NO_DISPONIBLE": (
        "Stock no disponible por warehouse y producto que debe descontarse del "
        "inventario disponible final."
    ),
    "STOCK": (
        "Stock disponible final por warehouse y producto. Es la fuente mandante "
        "para determinar cuánto puede enviarse."
    ),
    "INSUMOS": (
        "Cantidades de insumos calculadas automáticamente por Aleph para cada "
        "tienda. Solo se anexan al BulkCD_444 cuando la tienda ya recibe producto "
        "normal desde ese mismo origen."
    ),
    "GOLDEN_INFALTABLES": (
        "Productos Golden e Infaltables definidos por producto y ciudad, con "
        "prioridad y excepciones especiales de negocio."
    ),
    "TIENDA": (
        "Catálogo de warehouses con el nombre y la ciudad de cada tienda o nodo."
    ),
    "STORAGE": (
        "Condición de almacenamiento de cada producto, como room temperature, "
        "refrigerated o freezer."
    ),
}

DEMAND_RULE_LABELS = {
    "MOV_MINIMO_3": "ROQ POSITIVO",
    "HARDCODE_4_CERO_TOTAL": "FORECAST 0 · FORZADO A 4",
    "HARDCODE_3_INVENTARIO_MENOR_DEMANDA": "ROQ 0 · INVENTARIO MENOR A DEMANDA",
    "SIN_DEMANDA": "SIN RECOMENDACIÓN",
    "AVL_DOH": "COBERTURA AVL POR DOH",
}

BREAKDOWN_ORDER = (
    "CORTE POR CIUDAD BLOQUEADA",
    "CORTE POR PRODUCTO RACKEADO 444",
    "CORTE POR STOCK",
    "OK COMPLETO POR FOUNTAIN9",
    "OK PARCIAL - CORTE POR PRODUCTO RACKEADO 444",
    "OK PARCIAL - CORTE POR STOCK",
    "ENVIADOS PARA CUBRIR AVL",
    "SIN RECOMENDACIÓN",
    "CORTE POR RUTA DE COSTOS",
    "CORTE POR TIENDA CERRADA",
    "INSUMOS",
    "CORTE POR CAPACIDAD DE TIENDA",
    "CORTE POR CAPACIDAD DE TAREAS",
    "OK PARCIAL - CORTE POR CAPACIDAD DE TAREAS",
    "CORTE POR BLOQUEO REGIONAL",
    "OK PARCIAL - CORTE POR BLOQUEO REGIONAL",
    "ERROR DE DATOS",
)

INSUMOS_COLUMNS = [
    "WAREHOUSE_DESTINATION",
    "WAREHOUSE_SOURCE",
    "RETAIL_ID",
    "QUANTITY",
    "PLANNED_DATE",
    "ROUTE",
    "DELIVERY_PRIORITY",
]

CITY_DISPLAY_NAMES = {
    "CDMX": "Ciudad de México",
    "GDL": "Guadalajara",
    "MTY": "Monterrey",
}

DEFAULT_STORAGE = "Room Temperature"
MISSING_STORAGE_VALUES = {"", "UNKNOWN", "UNKNOW", "N/A", "NA", "NONE", "NULL"}

ENGLISH_MONTHS = {
    "JANUARY": 1,
    "FEBRUARY": 2,
    "MARCH": 3,
    "APRIL": 4,
    "MAY": 5,
    "JUNE": 6,
    "JULY": 7,
    "AUGUST": 8,
    "SEPTEMBER": 9,
    "OCTOBER": 10,
    "NOVEMBER": 11,
    "DECEMBER": 12,
}

TIMEZONE_OFFSETS = {
    "UTC": 0,
    "GMT": 0,
    "EDT": -4,
    "EST": -5,
    "CDT": -5,
    "CST": -6,
    "MDT": -6,
    "MST": -7,
    "PDT": -7,
    "PST": -8,
}


st.set_page_config(
    page_title=APP_NAME,
    page_icon="↗",
    layout="wide",
    initial_sidebar_state="collapsed",
)


def inject_styles() -> None:
    st.markdown(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Archivo+Black&family=IBM+Plex+Mono:wght@400;500;600;700&display=swap');

        :root {
            --ink: #111111;
            --paper: #f2efe6;
            --acid: #d9ff3f;
            --coral: #ff5a47;
            --blue: #5e7cff;
            --white: #fffdf7;
            --muted: #6f6b63;
        }

        html, body, [class*="css"] {
            font-family: "IBM Plex Mono", monospace;
        }

        .stApp {
            background:
                linear-gradient(rgba(17,17,17,.055) 1px, transparent 1px),
                linear-gradient(90deg, rgba(17,17,17,.055) 1px, transparent 1px),
                var(--paper);
            background-size: 28px 28px;
            color: var(--ink);
        }

        [data-testid="stHeader"] { background: transparent; }
        [data-testid="stToolbar"], #MainMenu, footer { visibility: hidden; }
        .block-container { max-width: 1440px; padding: 2rem 3rem 4rem; }

        .hero {
            border: 4px solid var(--ink);
            background: var(--coral);
            box-shadow: 10px 10px 0 var(--ink);
            padding: 28px 32px;
            margin: 6px 10px 34px 0;
            position: relative;
            overflow: hidden;
        }

        .hero::after {
            content: "RUN / ALLOCATE / EXPORT";
            position: absolute;
            right: -42px;
            top: 34px;
            transform: rotate(8deg);
            border: 3px solid var(--ink);
            background: var(--acid);
            padding: 8px 46px;
            font-weight: 800;
            letter-spacing: .08em;
        }

        .hero-kicker {
            display: inline-block;
            border: 3px solid var(--ink);
            background: var(--white);
            padding: 5px 9px;
            font-weight: 800;
            margin-bottom: 18px;
        }

        .hero h1 {
            font-family: "Archivo Black", sans-serif;
            font-size: clamp(2.6rem, 7vw, 6.5rem);
            line-height: .88;
            letter-spacing: -.06em;
            margin: 0;
            max-width: 1050px;
            color: var(--ink);
        }

        .hero p {
            max-width: 780px;
            font-weight: 700;
            font-size: 1rem;
            margin: 22px 0 0;
        }

        .section-label {
            display: inline-block;
            background: var(--ink);
            color: var(--white);
            border: 3px solid var(--ink);
            padding: 7px 12px;
            margin: 18px 0 10px;
            font-weight: 800;
            letter-spacing: .08em;
        }

        .info-strip {
            border: 3px solid var(--ink);
            background: var(--acid);
            box-shadow: 6px 6px 0 var(--ink);
            padding: 14px 16px;
            margin: 8px 7px 24px 0;
            font-weight: 700;
        }

        .database-status {
            border: 4px solid var(--ink);
            box-shadow: 8px 8px 0 var(--ink);
            padding: 20px 22px;
            margin: 8px 9px 22px 0;
        }

        .database-status.online { background: var(--acid); }
        .database-status.review { background: var(--coral); }

        .database-title {
            font-family: "Archivo Black", sans-serif;
            font-size: clamp(1.4rem, 3vw, 2.4rem);
            line-height: 1;
        }

        .source-card {
            height: 138px;
            border: 3px solid var(--ink);
            box-shadow: 5px 5px 0 var(--ink);
            padding: 14px 15px;
            margin: 0 5px 18px 0;
            box-sizing: border-box;
            display: flex;
            flex-direction: column;
            justify-content: space-between;
            position: relative;
            overflow: hidden;
        }

        .source-card.ok { background: #baf264; }
        .source-card.error { background: var(--coral); }

        .source-card-name {
            font-family: "Archivo Black", sans-serif;
            font-size: .95rem;
            line-height: 1.05;
            overflow-wrap: anywhere;
        }

        .source-card-meta {
            margin-top: 9px;
            color: #242424;
            font-size: .68rem;
            font-weight: 800;
        }

        .source-card-result {
            margin-top: 13px;
            border-top: 2px solid var(--ink);
            padding-top: 9px;
            font-size: .78rem;
            font-weight: 900;
            text-transform: uppercase;
        }

        .source-card-tooltip {
            position: absolute;
            inset: 0;
            z-index: 5;
            box-sizing: border-box;
            background: var(--ink);
            color: var(--white);
            padding: 12px 13px;
            opacity: 0;
            visibility: hidden;
            overflow-y: auto;
            font-size: .68rem;
            font-weight: 700;
            line-height: 1.35;
            text-transform: none;
            transition: opacity .12s ease, visibility 0s linear .12s;
        }

        .source-card-tooltip strong {
            display: block;
            color: var(--acid);
            margin-bottom: 6px;
            font-size: .67rem;
            letter-spacing: .05em;
        }

        .source-card:hover .source-card-tooltip {
            opacity: 1;
            visibility: visible;
            transition: opacity .12s ease 1s, visibility 0s linear 1s;
        }

        div[data-testid="stFileUploader"] {
            border: 4px dashed var(--ink);
            background: var(--white);
            box-shadow: 8px 8px 0 var(--ink);
            padding: 18px;
            margin: 8px 9px 24px 0;
        }

        div[data-testid="stFileUploaderDropzone"] {
            background: var(--white);
            border: 0;
            min-height: 190px;
        }

        div[data-testid="stForm"], div[data-testid="stExpander"] {
            border: 3px solid var(--ink);
            border-radius: 0;
            background: rgba(255,253,247,.92);
            box-shadow: 6px 6px 0 var(--ink);
        }

        div[data-testid="stTextInput"] input,
        div[data-testid="stNumberInput"] input,
        div[data-testid="stDateInput"] input,
        div[data-baseweb="select"] > div {
            border: 2px solid var(--ink) !important;
            border-radius: 0 !important;
            background: var(--white) !important;
        }

        .stButton > button,
        .stDownloadButton > button,
        div[data-testid="stFormSubmitButton"] > button {
            border: 3px solid var(--ink);
            border-radius: 0;
            background: var(--acid);
            color: var(--ink);
            box-shadow: 5px 5px 0 var(--ink);
            font-family: "IBM Plex Mono", monospace;
            font-weight: 900;
            text-transform: uppercase;
            min-height: 50px;
            transition: transform .08s ease, box-shadow .08s ease;
        }

        .stButton > button:hover,
        .stDownloadButton > button:hover,
        div[data-testid="stFormSubmitButton"] > button:hover {
            color: var(--ink);
            border-color: var(--ink);
            transform: translate(3px, 3px);
            box-shadow: 2px 2px 0 var(--ink);
        }

        div[data-testid="stMetric"] {
            border: 3px solid var(--ink);
            background: var(--white);
            box-shadow: 5px 5px 0 var(--ink);
            padding: 14px 16px;
        }

        div[data-testid="stMetricLabel"] { font-weight: 800; }
        div[data-testid="stMetricValue"] { font-family: "Archivo Black", sans-serif; }

        .kpi-card {
            min-height: 132px;
            border: 3px solid var(--ink);
            background: var(--white);
            box-shadow: 5px 5px 0 var(--ink);
            padding: 13px 15px;
            margin: 0 5px 18px 0;
            box-sizing: border-box;
            position: relative;
            overflow: hidden;
            display: flex;
            flex-direction: column;
            justify-content: space-between;
        }

        .kpi-card.acid { background: var(--acid); }
        .kpi-card.blue { background: var(--blue); color: var(--white); }
        .kpi-card.coral { background: var(--coral); }

        .kpi-card-category {
            font-size: .62rem;
            font-weight: 900;
            letter-spacing: .08em;
            opacity: .75;
        }

        .kpi-card-label {
            margin-top: 7px;
            font-size: .72rem;
            font-weight: 900;
            line-height: 1.2;
        }

        .kpi-card-value {
            margin-top: 10px;
            font-family: "Archivo Black", sans-serif;
            font-size: clamp(1.45rem, 2.6vw, 2.25rem);
            line-height: 1;
        }

        .kpi-card-tooltip {
            position: absolute;
            inset: 0;
            z-index: 5;
            box-sizing: border-box;
            background: var(--ink);
            color: var(--white);
            padding: 13px 14px;
            opacity: 0;
            visibility: hidden;
            overflow-y: auto;
            font-size: .69rem;
            font-weight: 700;
            line-height: 1.4;
            transition: opacity .12s ease, visibility 0s linear .12s;
        }

        .kpi-card-tooltip strong {
            display: block;
            color: var(--acid);
            margin-bottom: 7px;
            font-size: .66rem;
            letter-spacing: .06em;
        }

        .kpi-card:hover .kpi-card-tooltip {
            opacity: 1;
            visibility: visible;
            transition: opacity .12s ease 1s, visibility 0s linear 1s;
        }

        [data-testid="stAlert"] {
            border: 3px solid var(--ink);
            border-radius: 0;
        }

        .result-title {
            font-family: "Archivo Black", sans-serif;
            font-size: clamp(2rem, 5vw, 4.3rem);
            line-height: .95;
            margin: 38px 0 18px;
        }

        .report-title {
            font-family: "Archivo Black", sans-serif;
            font-size: clamp(1.8rem, 4vw, 3.1rem);
            line-height: .95;
            margin: 52px 0 18px;
        }

        .report-note {
            border: 3px solid var(--ink);
            background: var(--blue);
            color: var(--white);
            box-shadow: 5px 5px 0 var(--ink);
            padding: 12px 14px;
            margin: 0 6px 22px 0;
            font-size: .78rem;
            font-weight: 800;
        }

        .origin-banner {
            border: 3px solid var(--ink);
            background: var(--acid);
            box-shadow: 6px 6px 0 var(--ink);
            padding: 14px 17px;
            margin: 30px 7px 20px 0;
            font-family: "Archivo Black", sans-serif;
            font-size: clamp(1.05rem, 2.2vw, 1.7rem);
            line-height: 1.1;
        }

        .file-pill {
            border: 2px solid var(--ink);
            background: var(--blue);
            color: white;
            padding: 7px 10px;
            display: inline-block;
            font-weight: 700;
        }

        @media (max-width: 800px) {
            .block-container { padding: 1rem 1rem 3rem; }
            .hero { padding: 22px 18px; }
            .hero::after { display: none; }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def format_origin(warehouse_id: int) -> str:
    return f"{warehouse_id} - {ORIGIN_WAREHOUSES[warehouse_id]}"


def database_sheet_type(sheet_name: str) -> str:
    if sheet_name in ALEPH_SHEETS:
        return "ALEPH"
    if sheet_name in MANUAL_BACKEND_SHEETS:
        return "BACKEND"
    return "IMPORTRANGE"


def database_sheet_control(sheet_name: str) -> str:
    if sheet_name in ALEPH_SHEETS:
        return "C7"
    return "A1"


def render_kpi_cards(
    cards: list[dict[str, Any]],
    *,
    columns_count: int = 4,
) -> None:
    """Renderiza KPIs brutalistas con definición visible tras 1 s de hover."""
    if not cards:
        return
    for start in range(0, len(cards), columns_count):
        columns = st.columns(columns_count)
        for column, card in zip(columns, cards[start : start + columns_count]):
            tone = str(card.get("tone", ""))
            if tone not in {"acid", "blue", "coral"}:
                tone = ""
            class_name = f"kpi-card {tone}".strip()
            card_html = (
                f'<div class="{class_name}">'
                f'<div><div class="kpi-card-category">'
                f'{html.escape(str(card.get("category", "KPI")))}</div>'
                f'<div class="kpi-card-label">'
                f'{html.escape(str(card["label"]))}</div></div>'
                f'<div class="kpi-card-value">'
                f'{html.escape(str(card["value"]))}</div>'
                f'<div class="kpi-card-tooltip"><strong>QUÉ SIGNIFICA</strong>'
                f'{html.escape(str(card["description"]))}</div>'
                '</div>'
            )
            with column:
                st.markdown(card_html, unsafe_allow_html=True)


def contains_ref_error(value: Any) -> bool:
    return value is not None and "#REF!" in str(value).upper()


def parse_update_timestamp(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, datetime.min.time())
    else:
        raw = str(value or "").strip()
        if not raw:
            return None

        english = re.fullmatch(
            r"([A-Za-z]+)\s+(\d{1,2}),\s*(\d{4}),?\s*"
            r"(\d{1,2}):(\d{2})(?::(\d{2}))?\s*"
            r"(AM|PM)?(?:\s+([A-Za-z]{2,5}))?",
            raw,
            flags=re.IGNORECASE,
        )
        if english:
            month_name, day, year, hour, minute, second, meridiem, zone = english.groups()
            month = ENGLISH_MONTHS.get(month_name.upper())
            if month is None:
                return None
            hour_number = int(hour)
            if meridiem and hour_number <= 12:
                if meridiem.upper() == "PM" and hour_number < 12:
                    hour_number += 12
                elif meridiem.upper() == "AM" and hour_number == 12:
                    hour_number = 0
            if hour_number > 23:
                return None
            zone_name = (zone or "").upper()
            tzinfo = (
                timezone(timedelta(hours=TIMEZONE_OFFSETS[zone_name]))
                if zone_name in TIMEZONE_OFFSETS
                else ZoneInfo("America/Mexico_City")
            )
            try:
                return datetime(
                    int(year),
                    month,
                    int(day),
                    hour_number,
                    int(minute),
                    int(second or 0),
                    tzinfo=tzinfo,
                )
            except ValueError:
                return None

        try:
            parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        except ValueError:
            parsed = None
            for date_format in (
                "%d/%m/%Y %H:%M:%S",
                "%m/%d/%Y %H:%M:%S",
                "%d-%m-%Y %H:%M:%S",
                "%Y-%m-%d %H:%M:%S",
            ):
                try:
                    parsed = datetime.strptime(raw, date_format)
                    break
                except ValueError:
                    continue
            if parsed is None:
                return None

    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=ZoneInfo("America/Mexico_City"))
    return parsed


def relative_update_text(
    value: Any,
    now: datetime | None = None,
) -> tuple[str, bool]:
    parsed = parse_update_timestamp(value)
    if parsed is None:
        return "C7 SIN FECHA VÁLIDA", False

    reference = now or datetime.now(ZoneInfo("America/Mexico_City"))
    if reference.tzinfo is None:
        reference = reference.replace(tzinfo=ZoneInfo("America/Mexico_City"))
    seconds = (reference.astimezone(timezone.utc) - parsed.astimezone(timezone.utc)).total_seconds()
    if seconds < -300:
        return "C7 TIENE FECHA FUTURA", False
    seconds = max(seconds, 0)

    if seconds < 60:
        return "Hace menos de 1 min", True
    if seconds < 3600:
        minutes = int(seconds // 60)
        return f"Hace {minutes} min", True
    if seconds < 86400:
        hours = int(seconds // 3600)
        return f"Hace {hours} h", True
    days = int(seconds // 86400)
    return f"Hace {days} día" if days == 1 else f"Hace {days} días", True


def inspect_database(workbook_bytes: bytes) -> dict[str, Any]:
    """Valida presencia de hojas, actualizaciones Aleph y errores IMPORTRANGE."""
    rows: list[dict[str, Any]] = []
    formula_workbook = openpyxl.load_workbook(
        io.BytesIO(workbook_bytes), read_only=True, data_only=False
    )
    value_workbook = openpyxl.load_workbook(
        io.BytesIO(workbook_bytes), read_only=True, data_only=True
    )
    try:
        for sheet_name in REQUIRED_DATABASE_SHEETS:
            if sheet_name not in formula_workbook.sheetnames:
                rows.append(
                    {
                        "HOJA": sheet_name,
                        "TIPO": database_sheet_type(sheet_name),
                        "CONTROL": database_sheet_control(sheet_name),
                        "DETALLE": "HOJA NO ENCONTRADA",
                        "ESTADO": "ERROR",
                    }
                )
                continue

            formula_sheet = formula_workbook[sheet_name]
            value_sheet = value_workbook[sheet_name]
            if sheet_name in ALEPH_SHEETS:
                value = value_sheet["C7"].value
                if value is None:
                    value = formula_sheet["C7"].value
                detail, healthy = relative_update_text(value)
                if contains_ref_error(value):
                    healthy = False
                    detail = "#REF! DETECTADO EN C7"
            elif sheet_name in MANUAL_BACKEND_SHEETS:
                formula_value = formula_sheet["A1"].value
                displayed_value = value_sheet["A1"].value
                header = engine.normalize_header(
                    displayed_value if displayed_value is not None else formula_value
                )
                healthy = header == "WAREHOUSE_ID"
                detail = (
                    "BLOQUEO BACKEND DISPONIBLE"
                    if healthy
                    else "A1 DEBE SER WAREHOUSE_ID"
                )
            else:
                formula_value = formula_sheet["A1"].value
                displayed_value = value_sheet["A1"].value
                healthy = not (
                    contains_ref_error(formula_value)
                    or contains_ref_error(displayed_value)
                )
                detail = "SIN #REF!" if healthy else "#REF! DETECTADO"

            rows.append(
                {
                    "HOJA": sheet_name,
                    "TIPO": database_sheet_type(sheet_name),
                    "CONTROL": database_sheet_control(sheet_name),
                    "DETALLE": detail,
                    "ESTADO": "OK" if healthy else "ERROR",
                }
            )
    finally:
        formula_workbook.close()
        value_workbook.close()

    healthy_count = sum(row["ESTADO"] == "OK" for row in rows)
    return {
        "online": healthy_count == len(rows),
        "healthy_count": healthy_count,
        "total_count": len(rows),
        "error_count": len(rows) - healthy_count,
        "rows": rows,
    }


def extract_available_cities(workbook_bytes: bytes) -> dict[str, str]:
    workbook = openpyxl.load_workbook(
        io.BytesIO(workbook_bytes), read_only=True, data_only=True
    )
    try:
        labels: dict[str, str] = {}
        for row in engine.iter_sheet_records(
            workbook,
            "TIENDA",
            ["CITY"],
            ["CITY"],
        ):
            raw_city = engine.clean_text(row["CITY"])
            normalized_city = engine.normalize_city(raw_city)
            if not normalized_city:
                continue
            labels.setdefault(
                normalized_city,
                CITY_DISPLAY_NAMES.get(normalized_city, raw_city or normalized_city),
            )
    finally:
        workbook.close()

    preferred_order = {"CDMX": 0, "GDL": 1, "MTY": 2}
    return dict(
        sorted(
            labels.items(),
            key=lambda item: (
                preferred_order.get(item[0], 99),
                item[1].upper(),
            ),
        )
    )


def save_uploaded_file(uploaded_file, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    uploaded_file.seek(0)
    with destination.open("wb") as handle:
        shutil.copyfileobj(uploaded_file, handle, length=8 * 1024 * 1024)
    uploaded_file.seek(0)


@st.cache_data(ttl=300, show_spinner=False)
def fetch_public_database() -> bytes:
    """Exporta el Google Sheet público completo como XLSX y conserva 5 min de caché."""
    export_url = (
        "https://docs.google.com/spreadsheets/d/"
        f"{DATA_TRANSFERS_SPREADSHEET_ID}/export?format=xlsx"
    )
    request = urllib.request.Request(
        export_url,
        headers={"User-Agent": "Mozilla/5.0 TransferPlanner/1.0"},
    )
    try:
        with urllib.request.urlopen(request, timeout=300) as response:
            workbook_bytes = response.read()
    except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError) as exc:
        raise RuntimeError(
            "No pude consultar la base de datos. Confirma que el Google Sheet "
            "siga configurado como 'Cualquier persona con el enlace: Lector'."
        ) from exc

    if not workbook_bytes:
        raise RuntimeError("Google devolvió una base de datos vacía.")
    if not zipfile.is_zipfile(io.BytesIO(workbook_bytes)):
        raise RuntimeError(
            "Google no devolvió un archivo Excel válido. Revisa el acceso público."
        )
    return workbook_bytes


def save_database(workbook_bytes: bytes, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    with destination.open("wb") as handle:
        handle.write(workbook_bytes)


def load_closed_store_ids(database_path: Path) -> set[int]:
    """Carga el bloqueo permanente de tiendas definido en DATA_TRANSFERS."""
    closed_stores: set[int] = set()
    workbook = openpyxl.load_workbook(database_path, read_only=True, data_only=True)
    try:
        for record in engine.iter_sheet_records(
            workbook,
            "TIENDAS_CERRADAS",
            ["WAREHOUSE_ID"],
            ["WAREHOUSE_ID"],
        ):
            warehouse = engine.to_id(
                record["WAREHOUSE_ID"],
                "TIENDAS_CERRADAS.WAREHOUSE_ID",
                allow_none=True,
            )
            if warehouse is not None:
                closed_stores.add(warehouse)
    finally:
        workbook.close()
    return closed_stores


def load_avl_catalog_rows(
    database_path: Path,
) -> tuple[list[dict[str, Any]], list[str]]:
    """Carga todo el catálogo y consolida una ADU por combinación tienda-SKU."""
    consolidated: dict[tuple[int, int], float] = {}
    warnings: list[str] = []
    workbook = openpyxl.load_workbook(database_path, read_only=True, data_only=True)
    try:
        for record in engine.iter_sheet_records(
            workbook,
            "CATALOGO",
            ["WAREHOUSE_ID", "PRODUCT_ID", "ADU"],
            ["WAREHOUSE_ID", "PRODUCT_ID", "ADU"],
        ):
            destination = engine.to_id(
                record["WAREHOUSE_ID"],
                "CATALOGO.WAREHOUSE_ID",
                allow_none=True,
            )
            sku = engine.to_id(
                record["PRODUCT_ID"],
                "CATALOGO.PRODUCT_ID",
                allow_none=True,
            )
            adu = max(engine.to_float(record["ADU"], 0.0), 0.0)
            if destination is None or sku is None or adu <= 0:
                continue
            key = (destination, sku)
            if key in consolidated and not math.isclose(
                consolidated[key], adu, abs_tol=1e-9
            ):
                previous = consolidated[key]
                consolidated[key] = max(previous, adu)
                if len(warnings) < 20:
                    warnings.append(
                        f"CATALOGO: {key} tiene ADU {previous} y {adu}; "
                        f"se usó la mayor ({consolidated[key]})."
                    )
            else:
                consolidated[key] = adu
    finally:
        workbook.close()

    rows = [
        {
            "WAREHOUSE_DESTINATION": destination,
            "RETAIL_ID": sku,
            "ADU": adu,
        }
        for (destination, sku), adu in consolidated.items()
    ]
    return rows, warnings


def load_insumos_rows(
    database_path: Path,
    catalogs,
) -> tuple[list[dict[str, Any]], list[str]]:
    """Lee INSUMOS tal como lo entrega Aleph; no recalcula sus cantidades."""
    rows: list[dict[str, Any]] = []
    warnings: list[str] = []
    workbook = openpyxl.load_workbook(database_path, read_only=True, data_only=True)
    try:
        for sheet_row, record in enumerate(
            engine.iter_sheet_records(
                workbook,
                "INSUMOS",
                INSUMOS_COLUMNS,
                INSUMOS_COLUMNS,
            ),
            start=14,
        ):
            destination = engine.to_id(
                record["WAREHOUSE_DESTINATION"],
                f"INSUMOS fila {sheet_row}.WAREHOUSE_DESTINATION",
                allow_none=True,
            )
            source = engine.to_id(
                record["WAREHOUSE_SOURCE"],
                f"INSUMOS fila {sheet_row}.WAREHOUSE_SOURCE",
                allow_none=True,
            )
            sku = engine.to_id(
                record["RETAIL_ID"],
                f"INSUMOS fila {sheet_row}.RETAIL_ID",
                allow_none=True,
            )
            quantity_value = engine.to_float(record["QUANTITY"], 0.0)

            if destination is None or source is None or sku is None:
                warnings.append(
                    f"INSUMOS fila {sheet_row}: se omitió por tener identificadores vacíos."
                )
                continue
            if quantity_value <= 0:
                continue
            if not math.isclose(
                quantity_value,
                round(quantity_value),
                abs_tol=1e-6,
            ):
                raise ValueError(
                    f"INSUMOS fila {sheet_row}: QUANTITY debe ser entera; "
                    f"se recibió {quantity_value!r}."
                )
            if source != 444:
                warnings.append(
                    f"INSUMOS fila {sheet_row}: se omitió el origen {source}; "
                    "los insumos solo pueden salir del warehouse 444."
                )
                continue

            store = catalogs.stores.get(destination, {})
            rows.append(
                {
                    "WAREHOUSE_DESTINATION": destination,
                    "WAREHOUSE_SOURCE": 444,
                    "RETAIL_ID": sku,
                    "QUANTITY": int(round(quantity_value)),
                    "PLANNED_DATE": "",
                    "ROUTE": 1,
                    "DELIVERY_PRIORITY": 1,
                    "CITY": store.get("city", ""),
                    "STORAGE": resolved_storage(catalogs.storage.get(sku)),
                    "VALUE": catalogs.high_value.get(sku, "REGULAR"),
                }
            )
    finally:
        workbook.close()
    return rows, warnings


def append_insumos_to_bulk_444(
    local_files: list[Path],
    result,
    insumos_rows: list[dict[str, Any]],
    origins: tuple[int, ...],
    include_insumos: bool,
) -> dict[str, Any]:
    """Anexa insumos sin modificar stock, capacidad ni el contador de tareas."""
    summary = {
        "requested": include_insumos,
        "enabled": include_insumos and 444 in origins,
        "source_rows": len(insumos_rows),
        "eligible_stores": 0,
        "lines_added": 0,
        "units_added": 0,
        "stores_added": 0,
    }
    if not summary["enabled"]:
        return summary

    regular_444_rows = [
        row
        for row in result.allocation_rows
        if row["WAREHOUSE_SOURCE"] == 444 and row["QUANTITY"] > 0
    ]
    eligible_destinations = {
        row["WAREHOUSE_DESTINATION"] for row in regular_444_rows
    }
    summary["eligible_stores"] = len(eligible_destinations)
    selected = [
        row
        for row in insumos_rows
        if row["WAREHOUSE_DESTINATION"] in eligible_destinations
    ]
    if not selected:
        return summary

    bulk_path = next(
        (path for path in local_files if path.name.lower() == "bulkcd_444.csv"),
        None,
    )
    if bulk_path is None:
        raise RuntimeError(
            "Se encontraron insumos elegibles, pero no se generó BulkCD_444.csv."
        )

    engine.write_csv(
        bulk_path,
        regular_444_rows + selected,
        engine.OUTPUT_COLUMNS,
    )
    summary.update(
        {
            "lines_added": len(selected),
            "units_added": sum(row["QUANTITY"] for row in selected),
            "stores_added": len(
                {row["WAREHOUSE_DESTINATION"] for row in selected}
            ),
        }
    )
    return summary


def render_database_health(health: dict[str, Any]) -> None:
    online = health["online"]
    css_class = "online" if online else "review"
    status = "ONLINE" if online else "REVISAR"
    st.markdown(
        f"""
        <div class="database-status {css_class}">
            <div class="database-title">BASE DE DATOS — {status}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    rows = health["rows"]
    for start in range(0, len(rows), 4):
        columns = st.columns(4)
        for column, row in zip(columns, rows[start : start + 4]):
            card_class = "ok" if row["ESTADO"] == "OK" else "error"
            if row["TIPO"] == "ALEPH":
                card_meta = "ÚLTIMA ACTUALIZACIÓN"
            elif row["TIPO"] == "BACKEND":
                card_meta = "BLOQUEO PERMANENTE"
            else:
                card_meta = "CONEXIÓN IMPORTRANGE"
            show_result = row["TIPO"] == "ALEPH" or row["ESTADO"] == "ERROR"
            result_html = (
                f'<div class="source-card-result">{html.escape(row["DETALLE"])}</div>'
                if show_result
                else ""
            )
            description = SHEET_DESCRIPTIONS.get(
                row["HOJA"], "Fuente de información utilizada por el modelo de abasto."
            )
            card_html = (
                f'<div class="source-card {card_class}">'
                f'<div><div class="source-card-name">{html.escape(row["HOJA"])}</div>'
                f'<div class="source-card-meta">{card_meta}</div></div>'
                f'{result_html}'
                f'<div class="source-card-tooltip"><strong>QUÉ CONTIENE</strong>'
                f'{html.escape(description)}</div>'
                '</div>'
            )
            with column:
                st.markdown(card_html, unsafe_allow_html=True)

    if not online:
        st.error(
            f"Se detectaron {health['error_count']} hojas con problemas. "
            "Revísalas antes de ejecutar la planeación."
        )


def create_zip(paths: list[Path], destination: Path) -> Path:
    with zipfile.ZipFile(destination, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for path in paths:
            archive.write(path, arcname=path.name)
    return destination


def ordered_breakdown_rows(
    status_counts: dict[str, int] | Counter[str],
) -> list[dict[str, Any]]:
    order = {label: index for index, label in enumerate(BREAKDOWN_ORDER)}
    return [
        {"BREAKDOWN": status, "FILAS": int(count)}
        for status, count in sorted(
            status_counts.items(),
            key=lambda item: (order.get(item[0], len(order)), item[0]),
        )
        if count
    ]


def clear_previous_workspace() -> None:
    previous = st.session_state.pop("last_workspace", "")
    if not previous:
        return
    path = Path(previous)
    temp_root = Path(tempfile.gettempdir()).resolve()
    resolved = path.resolve()
    if temp_root in resolved.parents and resolved.name.startswith("transfer_planner_"):
        shutil.rmtree(resolved, ignore_errors=True)


def percentage(numerator: float, denominator: float) -> float:
    if denominator <= 0:
        return 0.0
    return round((numerator / denominator) * 100, 1)


def resolved_storage(value: Any) -> str:
    raw = engine.clean_text(value)
    if raw.upper() in MISSING_STORAGE_VALUES:
        return DEFAULT_STORAGE
    return raw


def is_fruver_storage(value: Any) -> bool:
    normalized = engine.normalize_header(value)
    return "FRUVER" in normalized


def normalize_result_storage(result) -> None:
    """Garantiza una clasificación operativa en reportes y archivos Bulk."""
    for row in result.base_rows:
        row["STORAGE"] = resolved_storage(row.get("STORAGE"))
    for row in result.allocation_rows:
        row["STORAGE"] = resolved_storage(row.get("STORAGE"))


def apply_fruver_811_block(
    catalogs,
    origins: tuple[int, ...],
    requested: bool,
) -> dict[str, Any]:
    """Vuelve no elegible el stock FRUVER del origen 811 antes de planear."""
    summary = {
        "requested": requested,
        "enabled": requested and 811 in origins,
        "products_identified": 0,
        "products_with_stock_blocked": 0,
        "units_blocked": 0.0,
    }
    if not summary["enabled"]:
        return summary

    fruver_skus = {
        sku
        for sku, storage_name in catalogs.storage.items()
        if is_fruver_storage(storage_name)
    }
    summary["products_identified"] = len(fruver_skus)
    for sku in fruver_skus:
        key = (811, sku)
        base_stock = max(float(catalogs.stock_base.get(key, 0.0)), 0.0)
        already_unavailable = max(
            float(catalogs.unavailable_stock.get(key, 0.0)),
            0.0,
        )
        newly_blocked = max(base_stock - already_unavailable, 0.0)
        if newly_blocked <= 0:
            continue
        catalogs.unavailable_stock[key] = max(already_unavailable, base_stock)
        summary["products_with_stock_blocked"] += 1
        summary["units_blocked"] += newly_blocked

    summary["units_blocked"] = round(summary["units_blocked"], 3)
    return summary


def build_planning_analytics(
    result,
    configured_origins: tuple[int, ...] = (),
) -> dict[str, Any]:
    base_rows = result.base_rows
    allocation_rows = result.allocation_rows

    def original_roq_units(row: dict[str, Any]) -> int:
        return max(int(math.ceil(max(float(row["MOV_ORIGINAL"]), 0.0))), 0)

    eligible_rows = [row for row in base_rows if row["CANTIDAD_OBJETIVO"] > 0]
    assigned_rows = [row for row in eligible_rows if row["CANTIDAD_ASIGNADA"] > 0]
    fully_covered_rows = [
        row
        for row in eligible_rows
        if row["CANTIDAD_ASIGNADA"] >= row["CANTIDAD_OBJETIVO"]
    ]
    partially_covered_rows = [
        row
        for row in eligible_rows
        if 0 < row["CANTIDAD_ASIGNADA"] < row["CANTIDAD_OBJETIVO"]
    ]
    not_assigned_rows = [
        row for row in eligible_rows if row["CANTIDAD_ASIGNADA"] <= 0
    ]

    city_accumulators: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "stores": set(),
            "products": set(),
            "units": 0,
            "tasks": 0,
            "m3": 0.0,
        }
    )
    store_accumulators: dict[tuple[str, int, str], dict[str, Any]] = defaultdict(
        lambda: {"products": set(), "units": 0, "tasks": 0, "m3": 0.0}
    )

    for row in assigned_rows:
        city = row.get("CITY") or "SIN CIUDAD"
        destination = row["WAREHOUSE_DESTINATION"]
        warehouse_name = row.get("WAREHOUSE_NAME") or "SIN NOMBRE"
        sku = row["RETAIL_ID"]
        units = int(row["CANTIDAD_ASIGNADA"])
        tasks = int(row["TAREAS_GENERADAS"])
        assigned_m3 = float(row["M3_ASIGNADO"])

        city_data = city_accumulators[city]
        city_data["stores"].add(destination)
        city_data["products"].add(sku)
        city_data["units"] += units
        city_data["tasks"] += tasks
        city_data["m3"] += assigned_m3

        store_data = store_accumulators[(city, destination, warehouse_name)]
        store_data["products"].add(sku)
        store_data["units"] += units
        store_data["tasks"] += tasks
        store_data["m3"] += assigned_m3

    city_rows = [
        {
            "CIUDAD": city,
            "TIENDAS_ATENDIDAS": len(data["stores"]),
            "PRODUCTOS_DISTINTOS": len(data["products"]),
            "UNIDADES": data["units"],
            "M3": round(data["m3"], 3),
            "TAREAS": data["tasks"],
        }
        for city, data in sorted(city_accumulators.items())
    ]
    store_rows = [
        {
            "CIUDAD": city,
            "WAREHOUSE_ID": destination,
            "TIENDA": warehouse_name,
            "PRODUCTOS_DISTINTOS": len(data["products"]),
            "UNIDADES": data["units"],
            "M3": round(data["m3"], 3),
            "TAREAS": data["tasks"],
        }
        for (city, destination, warehouse_name), data in store_accumulators.items()
    ]
    store_rows.sort(key=lambda row: (row["CIUDAD"], -row["UNIDADES"], row["WAREHOUSE_ID"]))

    stockout_rows = [
        row for row in eligible_rows if bool(row.get("ES_STOCKOUT", False))
    ]
    stockout_served = [row for row in stockout_rows if row["CANTIDAD_ASIGNADA"] > 0]
    stockout_fully_covered = [
        row
        for row in stockout_rows
        if row["CANTIDAD_ASIGNADA"] >= row["CANTIDAD_OBJETIVO"]
    ]
    forecast_zero_forced = [
        row
        for row in eligible_rows
        if row["REGLA_DEMANDA"] == "HARDCODE_4_CERO_TOTAL"
    ]
    forecast_zero_served = [
        row for row in forecast_zero_forced if row["CANTIDAD_ASIGNADA"] > 0
    ]
    deficit_forced = [
        row
        for row in eligible_rows
        if row["REGLA_DEMANDA"] == "HARDCODE_3_INVENTARIO_MENOR_DEMANDA"
    ]
    minimum_three_applied = [
        row
        for row in eligible_rows
        if 0 < row["MOV_ORIGINAL"] < 3 and row["CANTIDAD_OBJETIVO"] == 3
    ]
    golden_rows = [
        row for row in eligible_rows if bool(row.get("ES_GOLDEN_INFALTABLE", False))
    ]

    rule_accumulators: dict[str, dict[str, float]] = defaultdict(
        lambda: {
            "cases": 0,
            "served": 0,
            "full": 0,
            "original_roq_units": 0,
            "hardcode_target_units": 0,
            "hardcode_assigned_units": 0,
            "target_units": 0,
            "assigned_units": 0,
        }
    )
    for row in base_rows:
        rule = row["REGLA_DEMANDA"]
        data = rule_accumulators[rule]
        target = int(row["CANTIDAD_OBJETIVO"])
        assigned = int(row["CANTIDAD_ASIGNADA"])
        original_roq = original_roq_units(row)
        hardcode_target = max(target - original_roq, 0)
        hardcode_assigned = max(assigned - original_roq, 0)
        data["cases"] += 1
        data["served"] += int(assigned > 0)
        data["full"] += int(target > 0 and assigned >= target)
        data["original_roq_units"] += original_roq
        data["hardcode_target_units"] += hardcode_target
        data["hardcode_assigned_units"] += hardcode_assigned
        data["target_units"] += target
        data["assigned_units"] += assigned

    rule_order = {rule: index for index, rule in enumerate(DEMAND_RULE_LABELS)}
    rule_rows = []
    for rule, data in sorted(
        rule_accumulators.items(), key=lambda item: rule_order.get(item[0], 999)
    ):
        rule_rows.append(
            {
                "REGLA": DEMAND_RULE_LABELS.get(rule, rule),
                "CASOS": int(data["cases"]),
                "CASOS_CON_ENVIO": int(data["served"]),
                "COBERTURA_COMPLETA": int(data["full"]),
                "UNIDADES_ROQ_ORIGINAL": int(data["original_roq_units"]),
                "INCREMENTO_HARDCODE_OBJETIVO": int(
                    data["hardcode_target_units"]
                ),
                "INCREMENTO_HARDCODE_ENVIADO": int(
                    data["hardcode_assigned_units"]
                ),
                "UNIDADES_OBJETIVO": int(data["target_units"]),
                "UNIDADES_ASIGNADAS": int(data["assigned_units"]),
                "COMPLIANCE_UNIDADES_%": percentage(
                    data["assigned_units"], data["target_units"]
                ),
            }
        )

    stockout_city_accumulators: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "cases": 0,
            "served": 0,
            "full": 0,
            "products": set(),
            "units": 0,
        }
    )
    for row in stockout_rows:
        city = row.get("CITY") or "SIN CIUDAD"
        data = stockout_city_accumulators[city]
        assigned = int(row["CANTIDAD_ASIGNADA"])
        data["cases"] += 1
        if assigned > 0:
            data["served"] += 1
            data["products"].add(row["RETAIL_ID"])
            data["units"] += assigned
        if assigned >= row["CANTIDAD_OBJETIVO"]:
            data["full"] += 1

    stockout_city_rows = [
        {
            "CIUDAD": city,
            "CASOS_STOCKOUT": data["cases"],
            "CASOS_CON_ENVIO": data["served"],
            "COBERTURA_COMPLETA": data["full"],
            "PRODUCTOS_DISTINTOS_ATENDIDOS": len(data["products"]),
            "UNIDADES_ASIGNADAS": data["units"],
            "ATENCION_%": percentage(data["served"], data["cases"]),
        }
        for city, data in sorted(stockout_city_accumulators.items())
    ]

    m3_by_destination_sku = {
        (row["WAREHOUSE_DESTINATION"], row["RETAIL_ID"]): row["M3_POR_UNIDAD"]
        for row in base_rows
    }
    base_by_destination_sku = {
        (row["WAREHOUSE_DESTINATION"], row["RETAIL_ID"]): row
        for row in base_rows
    }
    source_accumulators: dict[int, dict[str, Any]] = defaultdict(
        lambda: {
            "tasks": 0,
            "units": 0,
            "products": set(),
            "stores": set(),
            "m3": 0.0,
        }
    )
    source_city_accumulators: dict[tuple[int, str], dict[str, Any]] = defaultdict(
        lambda: {
            "stores": set(),
            "products": set(),
            "tasks": 0,
            "units": 0,
            "m3": 0.0,
        }
    )
    source_store_accumulators: dict[
        tuple[int, str, int, str], dict[str, Any]
    ] = defaultdict(
        lambda: {"products": set(), "tasks": 0, "units": 0, "m3": 0.0}
    )
    source_storage_accumulators: dict[
        tuple[int, str], dict[str, Any]
    ] = defaultdict(
        lambda: {
            "products": set(),
            "stores": set(),
            "tasks": 0,
            "units": 0,
            "m3": 0.0,
        }
    )

    source_order = list(dict.fromkeys(configured_origins))
    for source in source_order:
        source_accumulators[source]

    for row in allocation_rows:
        source = row["WAREHOUSE_SOURCE"]
        if source not in source_order:
            source_order.append(source)
        data = source_accumulators[source]
        quantity = int(row["QUANTITY"])
        destination = row["WAREHOUSE_DESTINATION"]
        sku = row["RETAIL_ID"]
        base_row = base_by_destination_sku.get((destination, sku), {})
        city = row.get("CITY") or base_row.get("CITY") or "SIN CIUDAD"
        warehouse_name = (
            base_row.get("WAREHOUSE_NAME") or f"WAREHOUSE {destination}"
        )
        storage = resolved_storage(row.get("STORAGE"))
        line_m3 = quantity * float(
            m3_by_destination_sku.get((destination, sku), 0.0)
        )
        data["tasks"] += 1
        data["units"] += quantity
        data["products"].add(sku)
        data["stores"].add(destination)
        data["m3"] += line_m3

        city_data = source_city_accumulators[(source, city)]
        city_data["stores"].add(destination)
        city_data["products"].add(sku)
        city_data["tasks"] += 1
        city_data["units"] += quantity
        city_data["m3"] += line_m3

        store_data = source_store_accumulators[
            (source, city, destination, warehouse_name)
        ]
        store_data["products"].add(sku)
        store_data["tasks"] += 1
        store_data["units"] += quantity
        store_data["m3"] += line_m3

        storage_data = source_storage_accumulators[(source, storage)]
        storage_data["products"].add(sku)
        storage_data["stores"].add(destination)
        storage_data["tasks"] += 1
        storage_data["units"] += quantity
        storage_data["m3"] += line_m3

    total_source_tasks = sum(
        source_accumulators[source]["tasks"] for source in source_order
    )
    total_source_units = sum(
        source_accumulators[source]["units"] for source in source_order
    )
    source_rows: list[dict[str, Any]] = []
    for source in source_order:
        data = source_accumulators[source]
        source_rows.append(
            {
                "WAREHOUSE_SOURCE": source,
                "NOMBRE": ORIGIN_WAREHOUSES.get(source, "ORIGEN CONFIGURADO"),
                "TAREAS": data["tasks"],
                "UNIDADES": data["units"],
                "PRODUCTOS_DISTINTOS": len(data["products"]),
                "TIENDAS_ATENDIDAS": len(data["stores"]),
                "M3": round(data["m3"], 3),
                "UNIDADES_POR_TAREA": round(
                    data["units"] / data["tasks"], 2
                )
                if data["tasks"]
                else 0.0,
                "PARTICIPACION_TAREAS_%": percentage(
                    data["tasks"], total_source_tasks
                ),
                "PARTICIPACION_UNIDADES_%": percentage(
                    data["units"], total_source_units
                ),
            }
        )

    source_details: list[dict[str, Any]] = []
    for source, source_summary in zip(source_order, source_rows):
        city_detail_rows = [
            {
                "CIUDAD": city,
                "TIENDAS_ATENDIDAS": len(data["stores"]),
                "PRODUCTOS_DISTINTOS": len(data["products"]),
                "TAREAS": data["tasks"],
                "UNIDADES": data["units"],
                "M3": round(data["m3"], 3),
            }
            for (row_source, city), data in source_city_accumulators.items()
            if row_source == source
        ]
        city_detail_rows.sort(key=lambda row: (-row["UNIDADES"], row["CIUDAD"]))

        store_detail_rows = [
            {
                "CIUDAD": city,
                "WAREHOUSE_ID": destination,
                "TIENDA": warehouse_name,
                "PRODUCTOS_DISTINTOS": len(data["products"]),
                "TAREAS": data["tasks"],
                "UNIDADES": data["units"],
                "M3": round(data["m3"], 3),
            }
            for (
                row_source,
                city,
                destination,
                warehouse_name,
            ), data in source_store_accumulators.items()
            if row_source == source
        ]
        store_detail_rows.sort(
            key=lambda row: (-row["UNIDADES"], row["WAREHOUSE_ID"])
        )

        storage_detail_rows = [
            {
                "STORAGE": storage,
                "PRODUCTOS_DISTINTOS": len(data["products"]),
                "TIENDAS_ATENDIDAS": len(data["stores"]),
                "TAREAS": data["tasks"],
                "UNIDADES": data["units"],
                "M3": round(data["m3"], 3),
            }
            for (
                row_source,
                storage,
            ), data in source_storage_accumulators.items()
            if row_source == source
        ]
        storage_detail_rows.sort(
            key=lambda row: (-row["UNIDADES"], row["STORAGE"])
        )

        source_details.append(
            {
                "warehouse_source": source,
                "name": ORIGIN_WAREHOUSES.get(source, "ORIGEN CONFIGURADO"),
                "summary": source_summary,
                "city_rows": city_detail_rows,
                "store_rows": store_detail_rows,
                "storage_rows": storage_detail_rows,
            }
        )

    target_units = sum(int(row["CANTIDAD_OBJETIVO"]) for row in eligible_rows)
    assigned_units = sum(int(row["CANTIDAD_ASIGNADA"]) for row in eligible_rows)
    original_roq_total = sum(original_roq_units(row) for row in eligible_rows)
    original_roq_fulfilled = sum(
        min(int(row["CANTIDAD_ASIGNADA"]), original_roq_units(row))
        for row in eligible_rows
    )
    hardcode_target_units = sum(
        max(int(row["CANTIDAD_OBJETIVO"]) - original_roq_units(row), 0)
        for row in eligible_rows
    )
    hardcode_assigned_units = sum(
        max(int(row["CANTIDAD_ASIGNADA"]) - original_roq_units(row), 0)
        for row in eligible_rows
    )
    hardcode_cases = sum(
        int(row["CANTIDAD_OBJETIVO"]) > original_roq_units(row)
        for row in eligible_rows
    )
    return {
        "city_rows": city_rows,
        "store_rows": store_rows,
        "rule_rows": rule_rows,
        "stockout_city_rows": stockout_city_rows,
        "source_rows": source_rows,
        "source_details": source_details,
        "summary": {
            "eligible_cases": len(eligible_rows),
            "fully_covered_cases": len(fully_covered_rows),
            "partial_cases": len(partially_covered_rows),
            "not_assigned_cases": len(not_assigned_rows),
            "target_units": target_units,
            "assigned_units": assigned_units,
            "case_compliance_pct": percentage(
                len(fully_covered_rows), len(eligible_rows)
            ),
            "case_service_pct": percentage(len(assigned_rows), len(eligible_rows)),
            "unit_compliance_pct": percentage(assigned_units, target_units),
            "original_roq_units": original_roq_total,
            "original_roq_fulfilled_units": original_roq_fulfilled,
            "original_roq_compliance_pct": percentage(
                original_roq_fulfilled, original_roq_total
            ),
            "hardcode_cases": hardcode_cases,
            "hardcode_target_units": hardcode_target_units,
            "hardcode_assigned_units": hardcode_assigned_units,
            "m3_assigned": round(
                sum(float(row["M3_ASIGNADO"]) for row in assigned_rows), 3
            ),
            "cities_served": len(city_rows),
            "stores_served": len(store_rows),
            "products_served": len({row["RETAIL_ID"] for row in assigned_rows}),
            "stockout_cases": len(stockout_rows),
            "stockout_cases_served": len(stockout_served),
            "stockout_cases_full": len(stockout_fully_covered),
            "stockout_products_served": len(
                {row["RETAIL_ID"] for row in stockout_served}
            ),
            "stockout_stores_served": len(
                {row["WAREHOUSE_DESTINATION"] for row in stockout_served}
            ),
            "stockout_attention_pct": percentage(
                len(stockout_served), len(stockout_rows)
            ),
            "forecast_zero_forced_cases": len(forecast_zero_forced),
            "forecast_zero_served_cases": len(forecast_zero_served),
            "forecast_zero_target_units": sum(
                int(row["CANTIDAD_OBJETIVO"]) for row in forecast_zero_forced
            ),
            "forecast_zero_assigned_units": sum(
                int(row["CANTIDAD_ASIGNADA"]) for row in forecast_zero_forced
            ),
            "deficit_forced_cases": len(deficit_forced),
            "deficit_forced_served_cases": sum(
                row["CANTIDAD_ASIGNADA"] > 0 for row in deficit_forced
            ),
            "minimum_three_cases": len(minimum_three_applied),
            "golden_cases": len(golden_rows),
            "golden_served_cases": sum(
                row["CANTIDAD_ASIGNADA"] > 0 for row in golden_rows
            ),
        },
    }


def apply_reporting_labels(result) -> None:
    """Cambia únicamente etiquetas de salida; las reglas internas no se alteran."""
    for row in result.base_rows:
        cut_type = str(row.get("TIPO_DE_CORTE", ""))
        assigned = int(row.get("CANTIDAD_ASIGNADA", 0) or 0)
        target = int(row.get("CANTIDAD_OBJETIVO", 0) or 0)
        stock_excluded_by_rack_444 = max(
            float(row.get("STOCK_BASE_444", 0) or 0)
            - float(row.get("NO_DISPONIBLE_444", 0) or 0)
            - float(row.get("COPERNICO_NO_USABLE_444", 0) or 0),
            0.0,
        )
        row["STOCK_EXCLUIDO_RACKEADO_444"] = stock_excluded_by_rack_444
        if (
            bool(row.get("RACKEADO_444", False))
            and stock_excluded_by_rack_444 > 0
            and assigned < target
            and "CORTE POR STOCK" in cut_type
        ):
            if assigned > 0:
                row["TIPO_DE_CORTE"] = (
                    "OK PARCIAL - CORTE POR PRODUCTO RACKEADO 444"
                )
                row["DETALLE_MOTIVO"] = (
                    f"Asignadas {assigned} de {target}; el SKU está excluido al "
                    "100% del stock utilizable del warehouse 444 por RACKEADOS."
                )
            else:
                row["TIPO_DE_CORTE"] = "CORTE POR PRODUCTO RACKEADO 444"
                row["DETALLE_MOTIVO"] = (
                    "El SKU está excluido al 100% del stock utilizable del "
                    "warehouse 444 por RACKEADOS."
                )

        if row.get("TIPO_DE_CORTE") == "SIN DEMANDA":
            row["TIPO_DE_CORTE"] = "SIN RECOMENDACIÓN"

        cut_label_map = {
            "OK": "OK COMPLETO POR FOUNTAIN9",
            "SIN RUTA DE COSTOS": "CORTE POR RUTA DE COSTOS",
            "BLOQUEO REGIONAL": "CORTE POR BLOQUEO REGIONAL",
            "OK PARCIAL - BLOQUEO REGIONAL": (
                "OK PARCIAL - CORTE POR BLOQUEO REGIONAL"
            ),
        }
        current_cut = row.get("TIPO_DE_CORTE")
        if current_cut in cut_label_map:
            row["TIPO_DE_CORTE"] = cut_label_map[current_cut]

        detail = row.get("DETALLE_MOTIVO")
        if isinstance(detail, str):
            row["DETALLE_MOTIVO"] = detail.replace("MOV", "ROQ").replace(
                "Sin demanda", "Sin recomendación"
            )

        demand_rule = row.get("REGLA_DEMANDA")
        if isinstance(demand_rule, str):
            row["REGLA_DEMANDA"] = demand_rule.replace(
                "SIN_DEMANDA", "SIN_RECOMENDACION"
            ).replace("MOV", "ROQ")

        if "MOV_ORIGINAL" in row:
            renamed_row: dict[str, Any] = {}
            for key, value in row.items():
                renamed_row["ROQ_ORIGINAL" if key == "MOV_ORIGINAL" else key] = value
            row.clear()
            row.update(renamed_row)


def build_golden_analytics(result) -> dict[str, Any]:
    """Construye el reporte específico de Golden Infaltables."""
    golden_rows = [
        row
        for row in result.base_rows
        if bool(row.get("ES_GOLDEN_INFALTABLE", False))
        and int(row.get("CANTIDAD_OBJETIVO", 0) or 0) > 0
    ]
    city_data: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "cases": 0,
            "served": 0,
            "full": 0,
            "products": set(),
            "stores": set(),
            "target": 0,
            "assigned": 0,
        }
    )
    store_data: dict[tuple[str, int, str], dict[str, Any]] = defaultdict(
        lambda: {
            "cases": 0,
            "served": 0,
            "full": 0,
            "products": set(),
            "target": 0,
            "assigned": 0,
        }
    )
    origin_data: dict[int, dict[str, Any]] = defaultdict(
        lambda: {
            "cases": 0,
            "products": set(),
            "stores": set(),
            "units": 0,
        }
    )
    detail_rows: list[dict[str, Any]] = []
    for row in golden_rows:
        city = row.get("CITY") or "SIN CIUDAD"
        destination = int(row["WAREHOUSE_DESTINATION"])
        warehouse_name = row.get("WAREHOUSE_NAME") or "SIN NOMBRE"
        sku = int(row["RETAIL_ID"])
        sku_name = engine.clean_text(row.get("SKU_NAME"))
        target = int(row.get("CANTIDAD_OBJETIVO", 0) or 0)
        assigned = int(row.get("CANTIDAD_ASIGNADA", 0) or 0)

        city_bucket = city_data[city]
        city_bucket["cases"] += 1
        city_bucket["served"] += int(assigned > 0)
        city_bucket["full"] += int(assigned >= target)
        city_bucket["products"].add(sku)
        city_bucket["stores"].add(destination)
        city_bucket["target"] += target
        city_bucket["assigned"] += assigned

        store_bucket = store_data[(city, destination, warehouse_name)]
        store_bucket["cases"] += 1
        store_bucket["served"] += int(assigned > 0)
        store_bucket["full"] += int(assigned >= target)
        store_bucket["products"].add(sku)
        store_bucket["target"] += target
        store_bucket["assigned"] += assigned

        for key, value in row.items():
            if not key.startswith("ASIGNADO_"):
                continue
            source_text = key.removeprefix("ASIGNADO_")
            if not source_text.isdigit():
                continue
            source_units = int(value or 0)
            if source_units <= 0:
                continue
            source = int(source_text)
            origin_bucket = origin_data[source]
            origin_bucket["cases"] += 1
            origin_bucket["products"].add(sku)
            origin_bucket["stores"].add(destination)
            origin_bucket["units"] += source_units

        detail_rows.append(
            {
                "CIUDAD": city,
                "TIENDA": f"{destination} · {warehouse_name}",
                "SKU": f"{sku} · {sku_name}" if sku_name else str(sku),
                "OBJETIVO": target,
                "ASIGNADO": assigned,
                "FALTANTE": max(target - assigned, 0),
                "ORÍGENES": row.get("ORIGENES_USADOS", ""),
                "BREAKDOWN": row.get("TIPO_DE_CORTE", ""),
            }
        )

    city_rows = [
        {
            "CIUDAD": city,
            "CASOS": data["cases"],
            "CON_ENVÍO": data["served"],
            "COMPLETOS": data["full"],
            "TIENDAS": len(data["stores"]),
            "PRODUCTOS": len(data["products"]),
            "OBJETIVO": data["target"],
            "ASIGNADO": data["assigned"],
            "COMPLIANCE_%": percentage(data["assigned"], data["target"]),
        }
        for city, data in sorted(city_data.items())
    ]
    store_rows = [
        {
            "CIUDAD": city,
            "TIENDA": f"{destination} · {warehouse_name}",
            "CASOS": data["cases"],
            "CON_ENVÍO": data["served"],
            "COMPLETOS": data["full"],
            "PRODUCTOS": len(data["products"]),
            "OBJETIVO": data["target"],
            "ASIGNADO": data["assigned"],
            "FALTANTE": max(data["target"] - data["assigned"], 0),
            "COMPLIANCE_%": percentage(data["assigned"], data["target"]),
        }
        for (city, destination, warehouse_name), data in store_data.items()
    ]
    store_rows.sort(key=lambda row: (-row["FALTANTE"], row["TIENDA"]))
    origin_rows = [
        {
            "ORIGEN": source,
            "NOMBRE": ORIGIN_WAREHOUSES.get(source, "ORIGEN CONFIGURADO"),
            "CASOS_CON_APORTE": data["cases"],
            "PRODUCTOS": len(data["products"]),
            "TIENDAS": len(data["stores"]),
            "UNIDADES": data["units"],
        }
        for source, data in sorted(origin_data.items())
    ]
    detail_rows.sort(
        key=lambda row: (-row["FALTANTE"], row["CIUDAD"], row["TIENDA"], row["SKU"])
    )

    target_units = sum(int(row["CANTIDAD_OBJETIVO"]) for row in golden_rows)
    assigned_units = sum(int(row["CANTIDAD_ASIGNADA"]) for row in golden_rows)
    served_rows = [row for row in golden_rows if row["CANTIDAD_ASIGNADA"] > 0]
    full_rows = [
        row
        for row in golden_rows
        if row["CANTIDAD_ASIGNADA"] >= row["CANTIDAD_OBJETIVO"]
    ]
    partial_rows = [
        row
        for row in golden_rows
        if 0 < row["CANTIDAD_ASIGNADA"] < row["CANTIDAD_OBJETIVO"]
    ]
    return {
        "summary": {
            "cases": len(golden_rows),
            "served_cases": len(served_rows),
            "full_cases": len(full_rows),
            "partial_cases": len(partial_rows),
            "not_served_cases": len(golden_rows) - len(served_rows),
            "products": len({row["RETAIL_ID"] for row in golden_rows}),
            "stores": len({row["WAREHOUSE_DESTINATION"] for row in golden_rows}),
            "cities": len({row.get("CITY") or "SIN CIUDAD" for row in golden_rows}),
            "target_units": target_units,
            "assigned_units": assigned_units,
            "missing_units": max(target_units - assigned_units, 0),
            "case_compliance_pct": percentage(len(full_rows), len(golden_rows)),
            "unit_compliance_pct": percentage(assigned_units, target_units),
            "m3_assigned": round(
                sum(float(row.get("M3_ASIGNADO", 0) or 0) for row in golden_rows),
                3,
            ),
        },
        "city_rows": city_rows,
        "store_rows": store_rows,
        "origin_rows": origin_rows,
        "detail_rows": detail_rows,
    }


def split_plan_rows_by_closed_store(
    plan_rows: list[dict[str, Any]],
    closed_store_ids: set[int],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    active_rows: list[dict[str, Any]] = []
    blocked_rows: list[dict[str, Any]] = []
    for row in plan_rows:
        if row["WAREHOUSE_DESTINATION"] in closed_store_ids:
            blocked_rows.append(row)
        else:
            active_rows.append(row)
    return active_rows, blocked_rows


def closed_store_summary(
    blocked_rows: list[dict[str, Any]],
    closed_store_ids: set[int],
    config: engine.Config,
) -> dict[str, Any]:
    impacted_stores = {
        row["WAREHOUSE_DESTINATION"] for row in blocked_rows
    }
    products = {row["RETAIL_ID"] for row in blocked_rows}
    target_units = sum(
        engine.calculate_target_quantity(row, config)[0] for row in blocked_rows
    )
    return {
        "configured_stores": len(closed_store_ids),
        "requirements": len(blocked_rows),
        "stores": len(impacted_stores),
        "store_ids": sorted(impacted_stores),
        "products": len(products),
        "target_units": target_units,
    }


def split_plan_rows_by_blocked_city(
    plan_rows: list[dict[str, Any]],
    catalogs,
    blocked_cities: tuple[str, ...],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    blocked_city_set = set(blocked_cities)
    active_rows: list[dict[str, Any]] = []
    blocked_rows: list[dict[str, Any]] = []
    for row in plan_rows:
        store = catalogs.stores.get(row["WAREHOUSE_DESTINATION"], {})
        if store.get("city_norm", "") in blocked_city_set:
            blocked_rows.append(row)
        else:
            active_rows.append(row)
    return active_rows, blocked_rows


def city_block_summary(
    blocked_rows: list[dict[str, Any]],
    catalogs,
    config: engine.Config,
    blocked_cities: tuple[str, ...],
) -> dict[str, Any]:
    rows_by_city: Counter[str] = Counter()
    city_names: dict[str, str] = {}
    for store in catalogs.stores.values():
        city_norm = store.get("city_norm", "")
        if city_norm in blocked_cities and city_norm not in city_names:
            city_names[city_norm] = CITY_DISPLAY_NAMES.get(
                city_norm,
                store.get("city") or city_norm,
            )
    target_units = 0
    stores: set[int] = set()
    products: set[int] = set()
    for row in blocked_rows:
        destination = row["WAREHOUSE_DESTINATION"]
        store = catalogs.stores.get(destination, {})
        city_norm = store.get("city_norm", "")
        rows_by_city[city_norm] += 1
        stores.add(destination)
        products.add(row["RETAIL_ID"])
        target, _ = engine.calculate_target_quantity(row, config)
        target_units += target

    return {
        "cities": [
            {
                "code": city,
                "name": city_names.get(city, CITY_DISPLAY_NAMES.get(city, city)),
                "requirements": rows_by_city.get(city, 0),
            }
            for city in blocked_cities
        ],
        "requirements": len(blocked_rows),
        "stores": len(stores),
        "products": len(products),
        "target_units": target_units,
    }


def empty_avl_summary(enabled: bool, doh: float) -> dict[str, Any]:
    return {
        "enabled": enabled,
        "doh": float(doh),
        "catalog_rows": 0,
        "stockout_candidates": 0,
        "cases_sent": 0,
        "cases_full": 0,
        "cases_partial": 0,
        "tasks_added": 0,
        "units_added": 0,
        "m3_added": 0.0,
        "stores": 0,
        "products": 0,
        "task_slots_before": 0,
        "task_slots_after": 0,
        "skipped_closed_store": 0,
        "skipped_blocked_city": 0,
        "skipped_missing_stock": 0,
        "skipped_not_stockout": 0,
        "skipped_already_served": 0,
        "skipped_route_cost": 0,
        "skipped_capacity": 0,
        "skipped_no_source_stock": 0,
    }


def apply_avl_fill(
    result,
    catalog_rows: list[dict[str, Any]],
    catalogs,
    config: engine.Config,
    closed_store_ids: set[int],
    blocked_cities: tuple[str, ...],
    doh: float,
) -> dict[str, Any]:
    """Usa tareas remanentes para cubrir stockouts del catálogo por ADU x DOH."""
    summary = empty_avl_summary(True, doh)
    summary["catalog_rows"] = len(catalog_rows)
    summary["task_slots_before"] = max(config.max_tasks - result.tasks_used, 0)
    for base_row in result.base_rows:
        base_row.setdefault("ADU_CATALOGO", "")
        base_row.setdefault("DOH_AVL", "")
    if doh <= 0 or summary["task_slots_before"] <= 0:
        summary["task_slots_after"] = summary["task_slots_before"]
        return summary

    blocked_city_set = set(blocked_cities)
    assigned_keys = {
        (row["WAREHOUSE_DESTINATION"], row["RETAIL_ID"])
        for row in result.allocation_rows
        if int(row.get("QUANTITY", 0) or 0) > 0
    }
    base_row_by_key = {
        (row["WAREHOUSE_DESTINATION"], row["RETAIL_ID"]): row
        for row in result.base_rows
    }

    consumed_stock: Counter[tuple[int, int]] = Counter()
    for allocation in result.allocation_rows:
        consumed_stock[
            (allocation["WAREHOUSE_SOURCE"], allocation["RETAIL_ID"])
        ] += int(allocation["QUANTITY"])

    stock_info_cache: dict[tuple[int, int], dict[str, Any]] = {}
    stock_remaining: dict[tuple[int, int], int] = {}

    def get_source_stock(source: int, sku: int) -> tuple[dict[str, Any], int]:
        key = (source, sku)
        if key not in stock_info_cache:
            info = engine.source_stock_components(catalogs, source, sku)
            stock_info_cache[key] = info
            stock_remaining[key] = max(
                int(info["adjusted"]) - consumed_stock.get(key, 0),
                0,
            )
        return stock_info_cache[key], stock_remaining[key]

    capacity_by_store = {
        row["WAREHOUSE_DESTINATION"]: row for row in result.capacity_rows
    }

    candidates: list[dict[str, Any]] = []
    for catalog_row in catalog_rows:
        destination = catalog_row["WAREHOUSE_DESTINATION"]
        sku = catalog_row["RETAIL_ID"]
        key = (destination, sku)
        if destination in closed_store_ids:
            summary["skipped_closed_store"] += 1
            continue
        store = catalogs.stores.get(destination)
        if not store:
            continue
        city_norm = store.get("city_norm", "")
        if city_norm in blocked_city_set:
            summary["skipped_blocked_city"] += 1
            continue
        if key not in catalogs.stock_base:
            summary["skipped_missing_stock"] += 1
            continue
        if catalogs.stock_base[key] > 0:
            summary["skipped_not_stockout"] += 1
            continue
        summary["stockout_candidates"] += 1
        if key in assigned_keys:
            summary["skipped_already_served"] += 1
            continue
        if key in catalogs.route_cost_blocks:
            summary["skipped_route_cost"] += 1
            continue

        is_golden = bool(city_norm) and (
            sku,
            city_norm,
        ) in catalogs.golden_infaltables
        candidates.append(
            {
                "WAREHOUSE_DESTINATION": destination,
                "RETAIL_ID": sku,
                "ADU": float(catalog_row["ADU"]),
                "STORE": store,
                "IS_GOLDEN": is_golden,
                "PRIORITY": catalogs.store_priority.get(destination, 100),
            }
        )

    candidates.sort(
        key=lambda row: (
            0 if row["IS_GOLDEN"] else 1,
            row["PRIORITY"],
            row["WAREHOUSE_DESTINATION"],
            row["RETAIL_ID"],
        )
    )

    stores_sent: set[int] = set()
    products_sent: set[int] = set()
    next_order = len(result.base_rows) + 1
    for candidate in candidates:
        if result.tasks_used >= config.max_tasks:
            break

        destination = candidate["WAREHOUSE_DESTINATION"]
        sku = candidate["RETAIL_ID"]
        store = candidate["STORE"]
        city = store.get("city", "")
        city_norm = store.get("city_norm", "")
        is_golden = candidate["IS_GOLDEN"]
        m3_per_unit = catalogs.volume_m3.get(
            sku,
            config.default_m3_per_unit,
        )
        capacity = catalogs.store_capacity.get(
            destination,
            config.default_store_capacity_m3,
        )
        capacity_row = capacity_by_store.get(destination)
        if capacity_row is None:
            capacity_row = {
                "WAREHOUSE_DESTINATION": destination,
                "WAREHOUSE_NAME": store.get("warehouse_name", ""),
                "CITY": city,
                "CAPACIDAD_M3": capacity,
                "M3_CONTABILIZADO_CAPACIDAD": 0.0,
                "M3_TOTAL_ASIGNADO_INCLUYE_GOLDEN": 0.0,
                "CAPACIDAD_CERRADA": False,
                "CAPACIDAD_SUPERADA_POR_LINEA": False,
            }
            result.capacity_rows.append(capacity_row)
            capacity_by_store[destination] = capacity_row

        cap_before = float(capacity_row["M3_CONTABILIZADO_CAPACIDAD"])
        if not is_golden and bool(capacity_row["CAPACIDAD_CERRADA"]):
            summary["skipped_capacity"] += 1
            continue

        target = max(int(math.ceil(candidate["ADU"] * doh)), 3)
        origin_info: dict[int, dict[str, Any]] = {}
        origin_before: dict[int, int] = {}
        regional_blocks: dict[int, bool] = {}
        candidate_allocations: list[tuple[int, int]] = []
        remaining_target = target
        for source in config.origin_warehouses:
            info, available = get_source_stock(source, sku)
            origin_info[source] = info
            origin_before[source] = available
            regional_block = engine.is_regional_block(
                catalogs,
                source,
                destination,
                sku,
                city_norm,
                is_golden,
            )
            regional_blocks[source] = regional_block
            if regional_block:
                continue
            quantity = min(remaining_target, available)
            if quantity > 0:
                candidate_allocations.append((source, quantity))
                remaining_target -= quantity
            if remaining_target <= 0:
                break

        for source in config.origin_warehouses:
            if source in origin_info:
                continue
            info, available = get_source_stock(source, sku)
            origin_info[source] = info
            origin_before[source] = available
            regional_blocks[source] = engine.is_regional_block(
                catalogs,
                source,
                destination,
                sku,
                city_norm,
                is_golden,
            )

        available_task_slots = max(config.max_tasks - result.tasks_used, 0)
        allocations = candidate_allocations[:available_task_slots]
        assigned = sum(quantity for _, quantity in allocations)
        if assigned <= 0:
            summary["skipped_no_source_stock"] += 1
            continue

        task_before = result.tasks_used
        assigned_by_origin = {
            source: 0 for source in config.origin_warehouses
        }
        for source, quantity in allocations:
            stock_remaining[(source, sku)] -= quantity
            assigned_by_origin[source] += quantity
            result.tasks_used += 1
            result.allocation_rows.append(
                {
                    "WAREHOUSE_DESTINATION": destination,
                    "WAREHOUSE_SOURCE": source,
                    "RETAIL_ID": sku,
                    "QUANTITY": int(quantity),
                    "PLANNED_DATE": "",
                    "ROUTE": 1,
                    "DELIVERY_PRIORITY": 1,
                    "CITY": city,
                    "STORAGE": resolved_storage(catalogs.storage.get(sku)),
                    "VALUE": catalogs.high_value.get(sku, "REGULAR"),
                }
            )

        assigned_m3 = assigned * m3_per_unit
        capacity_row["M3_TOTAL_ASIGNADO_INCLUYE_GOLDEN"] = (
            float(capacity_row["M3_TOTAL_ASIGNADO_INCLUYE_GOLDEN"])
            + assigned_m3
        )
        if not is_golden:
            cap_after = cap_before + assigned_m3
            capacity_row["M3_CONTABILIZADO_CAPACIDAD"] = cap_after
            if cap_after >= capacity:
                capacity_row["CAPACIDAD_CERRADA"] = True
            if cap_after > capacity:
                capacity_row["CAPACIDAD_SUPERADA_POR_LINEA"] = True
        else:
            cap_after = cap_before

        previous_row = base_row_by_key.get((destination, sku))
        report_row: dict[str, Any] = {
            "ORDEN_PLANIFICACION": next_order,
            "FILA_INPUT": previous_row.get("FILA_INPUT", "CATALOGO")
            if previous_row
            else "CATALOGO",
            "FILAS_INPUT_CONSOLIDADAS": previous_row.get(
                "FILAS_INPUT_CONSOLIDADAS", "CATALOGO"
            )
            if previous_row
            else "CATALOGO",
            "CANTIDAD_FILAS_INPUT": previous_row.get("CANTIDAD_FILAS_INPUT", 0)
            if previous_row
            else 0,
            "DUPLICADO_CONFLICTIVO": False,
            "WAREHOUSE_DESTINATION": destination,
            "WAREHOUSE_NAME": store.get("warehouse_name", ""),
            "CITY": city,
            "RETAIL_ID": sku,
            "SKU_NAME": previous_row.get("SKU_NAME", "") if previous_row else "",
            "PREDICTED_OPENING_INVENTORY": previous_row.get(
                "PREDICTED_OPENING_INVENTORY", 0
            )
            if previous_row
            else 0,
            "PREDICTED_DEMAND": previous_row.get("PREDICTED_DEMAND", 0)
            if previous_row
            else 0,
            "CURRENT_INVENTORY": catalogs.stock_base.get((destination, sku), 0),
            "MOV_ORIGINAL": 0,
            "REGLA_DEMANDA": "AVL_DOH",
            "ADU_CATALOGO": candidate["ADU"],
            "DOH_AVL": doh,
            "CANTIDAD_OBJETIVO": target,
            "CANTIDAD_ASIGNADA": assigned,
            "CANTIDAD_FALTANTE": max(target - assigned, 0),
            "ES_GOLDEN_INFALTABLE": is_golden,
            "PRIORIDAD_TIENDA": candidate["PRIORITY"],
            "ES_STOCKOUT": True,
            "SIN_RUTA_COSTOS": False,
            "M3_POR_UNIDAD": m3_per_unit,
            "M3_OBJETIVO": target * m3_per_unit,
            "M3_ASIGNADO": assigned_m3,
            "CAPACIDAD_TIENDA_M3": capacity,
            "M3_CAPACIDAD_ANTES": cap_before,
            "M3_CAPACIDAD_DESPUES": cap_after,
            "EXCEDE_CAPACIDAD_EN_ESTA_LINEA": (
                not is_golden and cap_after > capacity
            ),
            "PASA_CAPACIDAD": True,
            "TAREAS_ANTES": task_before,
            "TAREAS_GENERADAS": len(allocations),
            "TAREAS_ACUMULADAS": result.tasks_used,
            "PASA_TAREAS": len(allocations) == len(candidate_allocations),
            "ORIGENES_USADOS": " | ".join(
                f"{source}:{quantity}" for source, quantity in allocations
            ),
            "STORAGE": resolved_storage(catalogs.storage.get(sku)),
            "VALUE": catalogs.high_value.get(sku, "REGULAR"),
            "TIPO_DE_CORTE": "ENVIADOS PARA CUBRIR AVL",
            "DETALLE_MOTIVO": (
                f"Stockout en destino; cobertura de {doh:g} DOH con "
                f"ADU {candidate['ADU']:.4f}. Asignadas {assigned} de {target}."
            ),
        }
        for source in config.origin_warehouses:
            info = origin_info[source]
            report_row.update(
                {
                    f"STOCK_BASE_{source}": info["base"],
                    f"NO_DISPONIBLE_{source}": info["unavailable"],
                    f"COPERNICO_NO_USABLE_{source}": info[
                        "copernico_unusable"
                    ],
                    f"RACKEADO_{source}": info["rackeado"],
                    f"STOCK_INICIAL_AJUSTADO_{source}": info["adjusted"],
                    f"STOCK_ANTES_{source}": origin_before[source],
                    f"BLOQUEO_REGIONAL_{source}": regional_blocks[source],
                    f"ASIGNADO_{source}": assigned_by_origin[source],
                    f"STOCK_REMANENTE_{source}": stock_remaining[(source, sku)],
                }
            )
        result.base_rows.append(report_row)
        base_row_by_key[(destination, sku)] = report_row
        next_order += 1
        summary["cases_sent"] += 1
        summary["cases_full"] += int(assigned >= target)
        summary["cases_partial"] += int(assigned < target)
        summary["tasks_added"] += len(allocations)
        summary["units_added"] += assigned
        summary["m3_added"] += assigned_m3
        stores_sent.add(destination)
        products_sent.add(sku)

    result.capacity_rows.sort(key=lambda row: row["WAREHOUSE_DESTINATION"])
    summary["m3_added"] = round(summary["m3_added"], 3)
    summary["stores"] = len(stores_sent)
    summary["products"] = len(products_sent)
    summary["task_slots_after"] = max(config.max_tasks - result.tasks_used, 0)
    return summary


def write_executive_pdf(
    path: Path,
    *,
    run_date: date,
    origins: tuple[int, ...],
    analytics: dict[str, Any],
    status_counts: dict[str, int] | Counter[str],
    input_requirements: int,
    evaluated_requirements: int,
    tasks: int,
    max_tasks: int,
    units: int,
    city_block: dict[str, Any],
    closed_stores: dict[str, Any],
    insumos: dict[str, Any],
    avl: dict[str, Any],
    fruver_811: dict[str, Any],
) -> None:
    """Genera un reporte PDF ejecutivo, legible y listo para compartir."""
    path.parent.mkdir(parents=True, exist_ok=True)
    page_width, page_height = landscape(A4)
    black = pdf_colors.HexColor("#111111")
    paper = pdf_colors.HexColor("#F6F2E8")
    acid = pdf_colors.HexColor("#CFFF2E")
    coral = pdf_colors.HexColor("#FF5B4D")
    blue = pdf_colors.HexColor("#5577FF")
    pale_green = pdf_colors.HexColor("#E6F7DF")
    pale_blue = pdf_colors.HexColor("#E9EEFF")
    grid = pdf_colors.HexColor("#D8D3C9")
    grey = pdf_colors.HexColor("#5A5A5A")

    base_styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "PdfTitle",
        parent=base_styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=25,
        leading=27,
        textColor=black,
        alignment=TA_LEFT,
        spaceAfter=3 * mm,
    )
    subtitle_style = ParagraphStyle(
        "PdfSubtitle",
        parent=base_styles["Normal"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=11,
        textColor=grey,
    )
    section_style = ParagraphStyle(
        "PdfSection",
        parent=base_styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=14,
        textColor=paper,
        backColor=black,
        borderPadding=(4, 7, 4, 7),
        spaceBefore=4 * mm,
        spaceAfter=2.5 * mm,
    )
    body_style = ParagraphStyle(
        "PdfBody",
        parent=base_styles["BodyText"],
        fontName="Helvetica",
        fontSize=8,
        leading=10.5,
        textColor=black,
    )
    small_style = ParagraphStyle(
        "PdfSmall",
        parent=body_style,
        fontSize=6.8,
        leading=8.4,
    )
    table_header_style = ParagraphStyle(
        "PdfTableHeader",
        parent=small_style,
        fontName="Helvetica-Bold",
        textColor=paper,
        alignment=TA_LEFT,
    )
    table_cell_style = ParagraphStyle(
        "PdfTableCell",
        parent=small_style,
        textColor=black,
    )
    card_label_style = ParagraphStyle(
        "PdfCardLabel",
        parent=small_style,
        fontName="Helvetica-Bold",
        fontSize=6.6,
        leading=8,
        textColor=black,
        alignment=TA_LEFT,
    )
    card_value_style = ParagraphStyle(
        "PdfCardValue",
        parent=body_style,
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=18,
        textColor=black,
        alignment=TA_LEFT,
    )

    def para(value: Any, style=table_cell_style) -> Paragraph:
        return Paragraph(html.escape(str(value)), style)

    def fmt_int(value: Any) -> str:
        return f"{int(value or 0):,}"

    def fmt_pct(value: Any) -> str:
        return f"{float(value or 0):,.1f}%"

    def fmt_m3(value: Any) -> str:
        return f"{float(value or 0):,.3f}"

    def report_table_pdf(
        rows: list[dict[str, Any]],
        columns: list[tuple[str, str, Any]],
        widths: list[float],
        *,
        max_rows: int | None = None,
    ) -> Table:
        selected = rows if max_rows is None else rows[:max_rows]
        data: list[list[Any]] = [
            [para(label, table_header_style) for _, label, _ in columns]
        ]
        for row in selected:
            data.append(
                [
                    para(formatter(row.get(key, "")), table_cell_style)
                    for key, _, formatter in columns
                ]
            )
        if not selected:
            data.append(
                [para("Sin registros", table_cell_style)]
                + [para("", table_cell_style) for _ in columns[1:]]
            )
        table = Table(data, colWidths=widths, repeatRows=1, hAlign="LEFT")
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), black),
                    ("TEXTCOLOR", (0, 0), (-1, 0), paper),
                    ("GRID", (0, 0), (-1, -1), 0.35, grid),
                    ("BACKGROUND", (0, 1), (-1, -1), pdf_colors.white),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 3.5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5),
                ]
            )
        )
        return table

    def kpi_cards(cards: list[tuple[str, str, Any]]) -> Table:
        cells = [
            [
                para(label.upper(), card_label_style),
                Spacer(1, 1.2 * mm),
                para(value, card_value_style),
            ]
            for label, value, _ in cards
        ]
        table = Table(
            [cells],
            colWidths=[(page_width - 28 * mm) / len(cells)] * len(cells),
        )
        style_commands: list[tuple[Any, ...]] = [
            ("BOX", (0, 0), (-1, -1), 1.2, black),
            ("INNERGRID", (0, 0), (-1, -1), 1.2, black),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]
        for column, (_, _, background) in enumerate(cards):
            style_commands.append(
                ("BACKGROUND", (column, 0), (column, 0), background)
            )
        table.setStyle(TableStyle(style_commands))
        return table

    summary = analytics["summary"]
    origin_text = " · ".join(
        f"{source} {ORIGIN_WAREHOUSES.get(source, '')}".strip()
        for source in origins
    )
    story: list[Any] = [
        Paragraph("REPORTE EJECUTIVO DE PLANEACIÓN", title_style),
        Paragraph(
            f"Fecha: {run_date:%d-%m-%Y} &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"Orígenes: {html.escape(origin_text)}",
            subtitle_style,
        ),
        Spacer(1, 4 * mm),
        kpi_cards(
            [
                ("Tareas de producto", fmt_int(tasks), acid),
                ("Unidades de producto", fmt_int(units), paper),
                ("Compliance de casos", fmt_pct(summary["case_compliance_pct"]), blue),
                ("Compliance de unidades", fmt_pct(summary["unit_compliance_pct"]), coral),
            ]
        ),
        Spacer(1, 3 * mm),
        kpi_cards(
            [
                ("Tiendas atendidas", fmt_int(summary["stores_served"]), paper),
                ("Productos distintos", fmt_int(summary["products_served"]), pale_blue),
                ("Volumen asignado m3", fmt_m3(summary["m3_assigned"]), pale_green),
                ("Uso del límite de tareas", f"{tasks:,} / {max_tasks:,}", paper),
            ]
        ),
        Paragraph("LECTURA EJECUTIVA", section_style),
    ]

    executive_rows = [
        [
            para("Demanda Fountain9", table_header_style),
            para("Stockouts", table_header_style),
            para("Cobertura AVL", table_header_style),
            para("Exclusiones y extras", table_header_style),
        ],
        [
            para(
                f"{input_requirements:,} casos únicos recibidos; "
                f"{evaluated_requirements:,} evaluados. "
                f"{summary['fully_covered_cases']:,} quedaron completos y "
                f"{summary['partial_cases']:,} parciales.",
                body_style,
            ),
            para(
                f"{summary['stockout_cases_served']:,} de "
                f"{summary['stockout_cases']:,} casos de stockout recibieron envío; "
                f"{summary['stockout_cases_full']:,} se cubrieron completamente.",
                body_style,
            ),
            para(
                (
                    f"Activada a {avl['doh']:g} DOH: "
                    f"{avl['cases_sent']:,} casos, {avl['tasks_added']:,} tareas y "
                    f"{avl['units_added']:,} unidades adicionales."
                )
                if avl.get("enabled")
                else "Cobertura AVL desactivada en esta ejecución.",
                body_style,
            ),
            para(
                f"{closed_stores.get('requirements', 0):,} casos por tienda cerrada; "
                f"{city_block.get('requirements', 0):,} por ciudad bloqueada; "
                f"{insumos.get('lines_added', 0):,} líneas y "
                f"{insumos.get('units_added', 0):,} unidades de insumos. "
                + (
                    f"FRUVER 811 bloqueado: "
                    f"{fruver_811.get('products_with_stock_blocked', 0):,} productos."
                    if fruver_811.get("enabled")
                    else "FRUVER 811 sin bloqueo manual."
                ),
                body_style,
            ),
        ],
    ]
    executive_table = Table(
        executive_rows,
        colWidths=[(page_width - 28 * mm) / 4] * 4,
    )
    executive_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), black),
                ("BACKGROUND", (0, 1), (-1, 1), pdf_colors.white),
                ("GRID", (0, 0), (-1, -1), 0.6, black),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.extend(
        [
            executive_table,
            Paragraph("BREAKDOWN DE LA PLANEACIÓN", section_style),
            report_table_pdf(
                ordered_breakdown_rows(status_counts),
                [
                    ("BREAKDOWN", "BREAKDOWN", str),
                    ("FILAS", "CASOS / LÍNEAS", fmt_int),
                ],
                [220 * mm, 45 * mm],
            ),
            PageBreak(),
            Paragraph("ANÁLISIS GENERAL", title_style),
            Paragraph("RESULTADO POR CIUDAD", section_style),
            report_table_pdf(
                analytics["city_rows"],
                [
                    ("CIUDAD", "CIUDAD", str),
                    ("TIENDAS_ATENDIDAS", "TIENDAS", fmt_int),
                    ("PRODUCTOS_DISTINTOS", "PRODUCTOS", fmt_int),
                    ("TAREAS", "TAREAS", fmt_int),
                    ("UNIDADES", "UNIDADES", fmt_int),
                    ("M3", "M3", fmt_m3),
                ],
                [80 * mm, 34 * mm, 38 * mm, 30 * mm, 37 * mm, 35 * mm],
            ),
            Paragraph("RESULTADO POR ORIGEN", section_style),
            report_table_pdf(
                analytics["source_rows"],
                [
                    ("WAREHOUSE_SOURCE", "ORIGEN", str),
                    ("NOMBRE", "NOMBRE", str),
                    ("TAREAS", "TAREAS", fmt_int),
                    ("UNIDADES", "UNIDADES", fmt_int),
                    ("PRODUCTOS_DISTINTOS", "PRODUCTOS", fmt_int),
                    ("TIENDAS_ATENDIDAS", "TIENDAS", fmt_int),
                    ("M3", "M3", fmt_m3),
                ],
                [25 * mm, 62 * mm, 28 * mm, 34 * mm, 33 * mm, 30 * mm, 32 * mm],
            ),
            Paragraph("TOP 15 TIENDAS POR UNIDADES", section_style),
            report_table_pdf(
                sorted(
                    analytics["store_rows"],
                    key=lambda row: (-row["UNIDADES"], row["WAREHOUSE_ID"]),
                ),
                [
                    ("CIUDAD", "CIUDAD", str),
                    ("WAREHOUSE_ID", "WH", str),
                    ("TIENDA", "TIENDA", str),
                    ("PRODUCTOS_DISTINTOS", "PRODUCTOS", fmt_int),
                    ("TAREAS", "TAREAS", fmt_int),
                    ("UNIDADES", "UNIDADES", fmt_int),
                    ("M3", "M3", fmt_m3),
                ],
                [48 * mm, 22 * mm, 75 * mm, 32 * mm, 26 * mm, 30 * mm, 28 * mm],
                max_rows=15,
            ),
        ]
    )

    golden = analytics.get("golden", {})
    golden_summary = golden.get("summary", {})
    if golden_summary:
        story.extend(
            [
                PageBreak(),
                Paragraph("GOLDEN INFALTABLES", title_style),
                Paragraph(
                    "Seguimiento exclusivo de los casos Golden: cobertura completa "
                    "por caso, cumplimiento de unidades y faltantes prioritarios.",
                    subtitle_style,
                ),
                Spacer(1, 3 * mm),
                kpi_cards(
                    [
                        ("Casos Golden", fmt_int(golden_summary["cases"]), blue),
                        ("Casos completos", fmt_int(golden_summary["full_cases"]), acid),
                        (
                            "Compliance de casos",
                            fmt_pct(golden_summary["case_compliance_pct"]),
                            paper,
                        ),
                        (
                            "Compliance de unidades",
                            fmt_pct(golden_summary["unit_compliance_pct"]),
                            coral,
                        ),
                    ]
                ),
                Paragraph("GOLDEN POR CIUDAD", section_style),
                report_table_pdf(
                    golden.get("city_rows", []),
                    [
                        ("CIUDAD", "CIUDAD", str),
                        ("CASOS", "CASOS", fmt_int),
                        ("COMPLETOS", "COMPLETOS", fmt_int),
                        ("TIENDAS", "TIENDAS", fmt_int),
                        ("OBJETIVO", "OBJETIVO", fmt_int),
                        ("ASIGNADO", "ASIGNADO", fmt_int),
                        ("COMPLIANCE_%", "COMPLIANCE", fmt_pct),
                    ],
                    [55 * mm, 27 * mm, 31 * mm, 28 * mm, 32 * mm, 32 * mm, 34 * mm],
                ),
                Paragraph("GOLDEN POR ORIGEN", section_style),
                report_table_pdf(
                    golden.get("origin_rows", []),
                    [
                        ("ORIGEN", "ORIGEN", str),
                        ("NOMBRE", "NOMBRE", str),
                        ("CASOS_CON_APORTE", "CASOS CON APORTE", fmt_int),
                        ("PRODUCTOS", "PRODUCTOS", fmt_int),
                        ("TIENDAS", "TIENDAS", fmt_int),
                        ("UNIDADES", "UNIDADES", fmt_int),
                    ],
                    [30 * mm, 85 * mm, 31 * mm, 35 * mm, 34 * mm, 38 * mm],
                ),
                Paragraph("TOP 15 TIENDAS GOLDEN POR FALTANTE", section_style),
                report_table_pdf(
                    golden.get("store_rows", []),
                    [
                        ("CIUDAD", "CIUDAD", str),
                        ("TIENDA", "TIENDA", str),
                        ("CASOS", "CASOS", fmt_int),
                        ("PRODUCTOS", "PRODUCTOS", fmt_int),
                        ("OBJETIVO", "OBJETIVO", fmt_int),
                        ("ASIGNADO", "ASIGNADO", fmt_int),
                        ("FALTANTE", "FALTANTE", fmt_int),
                        ("COMPLIANCE_%", "COMPLIANCE", fmt_pct),
                    ],
                    [34 * mm, 76 * mm, 22 * mm, 28 * mm, 27 * mm, 27 * mm, 26 * mm, 30 * mm],
                    max_rows=15,
                ),
                PageBreak(),
                Paragraph("DETALLE GOLDEN TIENDA-SKU", title_style),
                Paragraph(
                    "Casos ordenados por unidades faltantes para facilitar la "
                    "gestión de excepciones prioritarias.",
                    subtitle_style,
                ),
                Spacer(1, 3 * mm),
                report_table_pdf(
                    golden.get("detail_rows", []),
                    [
                        ("CIUDAD", "CIUDAD", str),
                        ("TIENDA", "TIENDA", str),
                        ("SKU", "SKU", str),
                        ("OBJETIVO", "OBJ.", fmt_int),
                        ("ASIGNADO", "ASIG.", fmt_int),
                        ("FALTANTE", "FALT.", fmt_int),
                        ("BREAKDOWN", "BREAKDOWN", str),
                    ],
                    [31 * mm, 59 * mm, 66 * mm, 21 * mm, 21 * mm, 21 * mm, 47 * mm],
                    max_rows=20,
                ),
            ]
        )

    for detail in analytics.get("source_details", []):
        source = detail["warehouse_source"]
        source_summary = detail["summary"]
        story.extend(
            [
                PageBreak(),
                Paragraph(
                    f"ANÁLISIS DEL ORIGEN {source}",
                    title_style,
                ),
                Paragraph(html.escape(detail["name"]), subtitle_style),
                Spacer(1, 3 * mm),
                kpi_cards(
                    [
                        ("Tareas", fmt_int(source_summary["TAREAS"]), acid),
                        ("Unidades", fmt_int(source_summary["UNIDADES"]), paper),
                        ("Tiendas", fmt_int(source_summary["TIENDAS_ATENDIDAS"]), blue),
                        ("Volumen m3", fmt_m3(source_summary["M3"]), coral),
                    ]
                ),
                Paragraph("DISTRIBUCIÓN POR CIUDAD", section_style),
                report_table_pdf(
                    detail["city_rows"],
                    [
                        ("CIUDAD", "CIUDAD", str),
                        ("TIENDAS_ATENDIDAS", "TIENDAS", fmt_int),
                        ("PRODUCTOS_DISTINTOS", "PRODUCTOS", fmt_int),
                        ("TAREAS", "TAREAS", fmt_int),
                        ("UNIDADES", "UNIDADES", fmt_int),
                        ("M3", "M3", fmt_m3),
                    ],
                    [76 * mm, 36 * mm, 40 * mm, 32 * mm, 38 * mm, 35 * mm],
                ),
                Paragraph("DISTRIBUCIÓN POR STORAGE", section_style),
                report_table_pdf(
                    detail["storage_rows"],
                    [
                        ("STORAGE", "STORAGE", str),
                        ("PRODUCTOS_DISTINTOS", "PRODUCTOS", fmt_int),
                        ("TIENDAS_ATENDIDAS", "TIENDAS", fmt_int),
                        ("TAREAS", "TAREAS", fmt_int),
                        ("UNIDADES", "UNIDADES", fmt_int),
                        ("M3", "M3", fmt_m3),
                    ],
                    [82 * mm, 40 * mm, 38 * mm, 32 * mm, 38 * mm, 34 * mm],
                ),
                Paragraph("TOP 12 TIENDAS DEL ORIGEN", section_style),
                report_table_pdf(
                    detail["store_rows"],
                    [
                        ("CIUDAD", "CIUDAD", str),
                        ("WAREHOUSE_ID", "WH", str),
                        ("TIENDA", "TIENDA", str),
                        ("PRODUCTOS_DISTINTOS", "PRODUCTOS", fmt_int),
                        ("TAREAS", "TAREAS", fmt_int),
                        ("UNIDADES", "UNIDADES", fmt_int),
                        ("M3", "M3", fmt_m3),
                    ],
                    [45 * mm, 22 * mm, 78 * mm, 34 * mm, 28 * mm, 30 * mm, 28 * mm],
                    max_rows=12,
                ),
            ]
        )

    def decorate_page(canvas, document) -> None:
        canvas.saveState()
        canvas.setFillColor(paper)
        canvas.rect(0, 0, page_width, page_height, fill=1, stroke=0)
        canvas.setFont("Helvetica", 6.5)
        canvas.setFillColor(grey)
        canvas.drawString(13 * mm, 6 * mm, f"Planeación {run_date:%d-%m-%Y}")
        canvas.drawRightString(
            page_width - 13 * mm,
            6 * mm,
            f"Página {document.page}",
        )
        canvas.restoreState()

    document = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        rightMargin=13 * mm,
        leftMargin=13 * mm,
        topMargin=13 * mm,
        bottomMargin=11 * mm,
        title=f"Reporte Ejecutivo de Planeación {run_date:%d-%m-%Y}",
        author="Transfer Planner",
        subject="Planeación ejecutiva de abasto y transferencias",
    )
    document.build(story, onFirstPage=decorate_page, onLaterPages=decorate_page)


def execute_planning(
    uploaded_plan,
    database_bytes: bytes,
    origins: tuple[int, ...],
    max_tasks: int,
    run_date,
    blocked_cities: tuple[str, ...] = (),
    include_insumos: bool = True,
    include_avl_fill: bool = False,
    avl_doh: float = 3.0,
    block_fruver_811: bool = False,
) -> dict[str, Any]:
    clear_previous_workspace()
    workspace = Path(tempfile.mkdtemp(prefix="transfer_planner_"))
    st.session_state["last_workspace"] = str(workspace)
    plan_path = workspace / "input" / Path(uploaded_plan.name).name
    data_path = workspace / "input" / "DATA_TRANSFERS.xlsx"

    save_uploaded_file(uploaded_plan, plan_path)
    save_database(database_bytes, data_path)

    config = engine.Config(
        origin_warehouses=origins,
        max_tasks=max_tasks,
        run_date_override=run_date.strftime("%d-%m-%Y"),
        default_store_capacity_m3=engine.CONFIG.default_store_capacity_m3,
        default_m3_per_unit=engine.CONFIG.default_m3_per_unit,
        minimum_positive_quantity=engine.CONFIG.minimum_positive_quantity,
        local_work_dir=str(workspace / "engine"),
        replace_same_day_outputs=True,
        generate_empty_source_files=False,
    )

    captured = io.StringIO()
    with redirect_stdout(captured):
        catalogs = engine.load_catalogs(data_path, config)
        fruver_811_summary = apply_fruver_811_block(
            catalogs,
            origins,
            block_fruver_811,
        )
        if fruver_811_summary["enabled"]:
            catalogs.warnings.append(
                "Bloqueo FRUVER 811: se volvió no elegible el stock de "
                f"{fruver_811_summary['products_with_stock_blocked']:,} productos "
                f"({fruver_811_summary['units_blocked']:,.0f} unidades) antes de "
                "la asignación. Los demás orígenes permanecieron disponibles."
            )
        closed_store_ids = load_closed_store_ids(data_path)
        insumos_rows: list[dict[str, Any]] = []
        if include_insumos and 444 in origins:
            insumos_rows, insumos_warnings = load_insumos_rows(
                data_path,
                catalogs,
            )
            catalogs.warnings.extend(insumos_warnings)
        plan_read = engine.read_plan_csv(plan_path, config)
        rows_after_closed_stores, closed_plan_rows = (
            split_plan_rows_by_closed_store(
                plan_read.rows,
                closed_store_ids,
            )
        )
        active_plan_rows, blocked_plan_rows = split_plan_rows_by_blocked_city(
            rows_after_closed_stores,
            catalogs,
            blocked_cities,
        )

        result = engine.plan_transfers(active_plan_rows, catalogs, config)
        result.warnings.extend(plan_read.warnings)
        closed_summary = closed_store_summary(
            closed_plan_rows,
            closed_store_ids,
            config,
        )
        if closed_plan_rows:
            closed_ids_text = ", ".join(map(str, closed_summary["store_ids"]))
            result.warnings.append(
                "Bloqueo permanente TIENDAS_CERRADAS: se excluyeron "
                f"{len(closed_plan_rows):,} requerimientos de las tiendas "
                f"{closed_ids_text} antes de asignar stock."
            )
        block_summary = city_block_summary(
            blocked_plan_rows,
            catalogs,
            config,
            blocked_cities,
        )
        if blocked_plan_rows:
            blocked_names = ", ".join(
                item["name"] for item in block_summary["cities"]
            )
            result.warnings.append(
                f"Bloqueo manual de ciudad: {blocked_names}. Se excluyeron "
                f"{len(blocked_plan_rows):,} requerimientos antes de asignar stock."
            )

        avl_summary = empty_avl_summary(include_avl_fill, avl_doh)
        if include_avl_fill:
            avl_catalog_rows, avl_warnings = load_avl_catalog_rows(data_path)
            result.warnings.extend(avl_warnings)
            avl_summary = apply_avl_fill(
                result,
                avl_catalog_rows,
                catalogs,
                config,
                closed_store_ids,
                blocked_cities,
                avl_doh,
            )
            result.warnings.append(
                f"Cobertura AVL ({avl_doh:g} DOH): se agregaron "
                f"{avl_summary['cases_sent']:,} casos, "
                f"{avl_summary['tasks_added']:,} tareas y "
                f"{avl_summary['units_added']:,} unidades usando exclusivamente "
                "tareas remanentes."
            )

        normalize_result_storage(result)
        analytics = build_planning_analytics(result, origins)
        apply_reporting_labels(result)
        analytics["golden"] = build_golden_analytics(result)
        output_dir = (
            Path(config.local_work_dir)
            / "outputs"
            / run_date.strftime("%d-%m-%Y")
        )
        local_files = engine.create_output_files(
            result,
            config,
            run_date,
            plan_path.name,
            output_dir,
        )
        local_files = [Path(path) for path in local_files]
        insumos_summary = append_insumos_to_bulk_444(
            local_files,
            result,
            insumos_rows,
            origins,
            include_insumos,
        )

        status_counts = Counter(
            row["TIPO_DE_CORTE"] for row in result.base_rows
        )
        if closed_summary["requirements"]:
            status_counts["CORTE POR TIENDA CERRADA"] += closed_summary[
                "requirements"
            ]
        if block_summary["requirements"]:
            status_counts["CORTE POR CIUDAD BLOQUEADA"] += block_summary[
                "requirements"
            ]
        if insumos_summary["lines_added"]:
            status_counts["INSUMOS"] += insumos_summary["lines_added"]

        units = sum(row["QUANTITY"] for row in result.allocation_rows)
        requirements = len(result.base_rows)
        pdf_path = output_dir / (
            f"Reporte_Ejecutivo_Planeacion_{run_date:%d-%m-%Y}.pdf"
        )
        write_executive_pdf(
            pdf_path,
            run_date=run_date,
            origins=origins,
            analytics=analytics,
            status_counts=status_counts,
            input_requirements=len(plan_read.rows),
            evaluated_requirements=requirements,
            tasks=result.tasks_used,
            max_tasks=config.max_tasks,
            units=units,
            city_block=block_summary,
            closed_stores=closed_summary,
            insumos=insumos_summary,
            avl=avl_summary,
            fruver_811=fruver_811_summary,
        )
        local_files.append(pdf_path)
        print(f"Requerimientos únicos del input: {len(plan_read.rows):,}")
        print(f"Requerimientos activos: {len(active_plan_rows):,}")
        print(
            "Requerimientos excluidos por tiendas cerradas: "
            f"{len(closed_plan_rows):,}"
        )
        print(f"Requerimientos excluidos por ciudad: {len(blocked_plan_rows):,}")
        print(f"Tareas generadas: {result.tasks_used:,}")
        if avl_summary["enabled"]:
            print(
                f"Cobertura AVL: {avl_summary['cases_sent']:,} casos / "
                f"{avl_summary['tasks_added']:,} tareas / "
                f"{avl_summary['units_added']:,} unidades"
            )
        if insumos_summary["enabled"]:
            print(
                "Insumos agregados a BulkCD_444: "
                f"{insumos_summary['lines_added']:,} líneas / "
                f"{insumos_summary['units_added']:,} unidades / "
                f"{insumos_summary['stores_added']:,} tiendas"
            )

    zip_path = create_zip(
        local_files,
        workspace / f"Planeacion_{run_date:%d-%m-%Y}.zip",
    )
    return {
        "workspace": str(workspace),
        "files": [str(path) for path in local_files],
        "zip": str(zip_path),
        "tasks": result.tasks_used,
        "units": units,
        "requirements": requirements,
        "input_requirements": len(plan_read.rows),
        "status_counts": dict(status_counts),
        "warnings": list(result.warnings),
        "logs": captured.getvalue(),
        "origins": list(origins),
        "analytics": analytics,
        "city_block": block_summary,
        "closed_stores": closed_summary,
        "insumos": insumos_summary,
        "avl": avl_summary,
        "fruver_811": fruver_811_summary,
    }


def report_table(
    rows: list[dict[str, Any]],
    *,
    column_config: dict[str, Any] | None = None,
    max_height: int = 560,
) -> None:
    if not rows:
        st.info("No existen registros para mostrar en esta sección.")
        return
    height = min(max(150, 36 * (len(rows) + 1)), max_height)
    st.dataframe(
        rows,
        use_container_width=True,
        hide_index=True,
        height=height,
        column_config=column_config or {},
    )


def render_source_analysis(analytics: dict[str, Any]) -> None:
    source_details = analytics.get("source_details", [])
    if not source_details:
        return

    st.markdown(
        '<div class="report-title">ANÁLISIS POR ORIGEN.</div>',
        unsafe_allow_html=True,
    )
    st.markdown(
        """
        <div class="report-note">
            Cada bloque corresponde a un warehouse origen seleccionado. TAREAS son
            líneas operativas del Bulk; UNIDADES son la suma de QUANTITY; PRODUCTOS son
            SKUs distintos. Este análisis considera únicamente abasto normal y excluye
            insumos para mantener comparables los orígenes. Mantén el cursor un segundo
            sobre cualquier tarjeta para consultar su definición.
        </div>
        """,
        unsafe_allow_html=True,
    )

    for detail in source_details:
        source = detail["warehouse_source"]
        name = detail["name"]
        summary = detail["summary"]
        st.markdown(
            f'<div class="origin-banner">{source} — {html.escape(str(name))}</div>',
            unsafe_allow_html=True,
        )
        render_kpi_cards(
            [
                {
                    "category": "TAREAS · ORIGEN",
                    "label": "LÍNEAS DE ABASTO",
                    "value": f"{summary['TAREAS']:,}",
                    "description": (
                        "Número de filas de transferencia generadas desde este origen. "
                        "Una misma tienda–SKU puede crear una tarea en más de un origen. "
                        "No incluye líneas de insumos."
                    ),
                    "tone": "blue",
                },
                {
                    "category": "UNIDADES · ORIGEN",
                    "label": "PRODUCTO ASIGNADO",
                    "value": f"{summary['UNIDADES']:,}",
                    "description": (
                        "Suma de QUANTITY de producto normal que saldrá desde este "
                        "warehouse origen. No incluye insumos."
                    ),
                    "tone": "acid",
                },
                {
                    "category": "SKUs · ORIGEN",
                    "label": "PRODUCTOS DISTINTOS",
                    "value": f"{summary['PRODUCTOS_DISTINTOS']:,}",
                    "description": (
                        "Cantidad de RETAIL_ID diferentes asignados desde este origen. "
                        "No es un conteo de tareas ni de unidades."
                    ),
                },
                {
                    "category": "TIENDAS · ORIGEN",
                    "label": "DESTINOS ATENDIDOS",
                    "value": f"{summary['TIENDAS_ATENDIDAS']:,}",
                    "description": (
                        "Número de warehouses destino diferentes que reciben producto "
                        "normal desde este origen."
                    ),
                },
                {
                    "category": "M³ · ORIGEN",
                    "label": "VOLUMEN PLANEADO",
                    "value": f"{summary['M3']:,.3f}",
                    "description": (
                        "Volumen de producto normal asignado desde este origen: unidades "
                        "por metros cúbicos por unidad."
                    ),
                },
                {
                    "category": "UNIDADES / TAREA",
                    "label": "PROMEDIO POR LÍNEA",
                    "value": f"{summary['UNIDADES_POR_TAREA']:,.2f}",
                    "description": (
                        "Unidades asignadas desde este origen divididas entre sus tareas. "
                        "Sirve para entender el tamaño promedio de cada línea operativa."
                    ),
                },
                {
                    "category": "PARTICIPACIÓN · TAREAS",
                    "label": "PESO OPERATIVO",
                    "value": f"{summary['PARTICIPACION_TAREAS_%']:,.1f}%",
                    "description": (
                        "Porcentaje de todas las tareas de abasto que salen desde este "
                        "origen. No utiliza unidades como denominador."
                    ),
                    "tone": "blue",
                },
                {
                    "category": "PARTICIPACIÓN · UNIDADES",
                    "label": "PESO EN UNIDADES",
                    "value": f"{summary['PARTICIPACION_UNIDADES_%']:,.1f}%",
                    "description": (
                        "Porcentaje de todas las unidades de abasto asignadas a este "
                        "origen. No utiliza tareas como denominador."
                    ),
                    "tone": "acid",
                },
            ]
        )

        left, right = st.columns(2)
        with left:
            st.markdown(
                '<span class="section-label">POR CIUDAD</span>',
                unsafe_allow_html=True,
            )
            compact_city_rows = [
                {
                    "CIUDAD": row["CIUDAD"],
                    "TAREAS": row["TAREAS"],
                    "UNIDADES": row["UNIDADES"],
                    "M3": row["M3"],
                }
                for row in detail["city_rows"]
            ]
            report_table(
                compact_city_rows,
                column_config={
                    "CIUDAD": st.column_config.TextColumn("CIUDAD", width="medium"),
                    "TAREAS": st.column_config.NumberColumn(
                        "TAREAS", format="%d", width="small"
                    ),
                    "UNIDADES": st.column_config.NumberColumn(
                        "UNIDADES", format="%d", width="small"
                    ),
                    "M3": st.column_config.NumberColumn(
                        "M³", format="%.3f", width="small"
                    ),
                },
                max_height=340,
            )
        with right:
            st.markdown(
                '<span class="section-label">POR STORAGE</span>',
                unsafe_allow_html=True,
            )
            compact_storage_rows = [
                {
                    "STORAGE": resolved_storage(row["STORAGE"]),
                    "TAREAS": row["TAREAS"],
                    "UNIDADES": row["UNIDADES"],
                    "M3": row["M3"],
                }
                for row in detail["storage_rows"]
            ]
            report_table(
                compact_storage_rows,
                column_config={
                    "STORAGE": st.column_config.TextColumn(
                        "STORAGE", width="medium"
                    ),
                    "TAREAS": st.column_config.NumberColumn(
                        "TAREAS", format="%d", width="small"
                    ),
                    "UNIDADES": st.column_config.NumberColumn(
                        "UNIDADES", format="%d", width="small"
                    ),
                    "M3": st.column_config.NumberColumn(
                        "M³", format="%.3f", width="small"
                    ),
                },
                max_height=340,
            )

        st.markdown(
            '<span class="section-label">TIENDAS ATENDIDAS DESDE ESTE ORIGEN</span>',
            unsafe_allow_html=True,
        )
        report_table(
            detail["store_rows"],
            column_config={
                "WAREHOUSE_ID": st.column_config.NumberColumn(
                    "WAREHOUSE ID", format="%d"
                ),
                "M3": st.column_config.NumberColumn("M³", format="%.3f"),
            },
            max_height=520,
        )


def render_golden_report(analytics: dict[str, Any]) -> None:
    golden = analytics.get("golden", {})
    summary = golden.get("summary", {})
    if not summary:
        return

    st.markdown(
        '<div class="report-title">GOLDEN INFALTABLES.</div>',
        unsafe_allow_html=True,
    )
    st.markdown(
        """
        <div class="report-note">
            Este bloque considera únicamente casos marcados como Golden Infaltables.
            COMPLIANCE DE CASOS exige cubrir el objetivo completo de cada tienda-SKU;
            COMPLIANCE DE UNIDADES compara las unidades asignadas contra las unidades
            objetivo. Las tarjetas distinguen casos, unidades, tiendas y productos.
        </div>
        """,
        unsafe_allow_html=True,
    )
    render_kpi_cards(
        [
            {
                "category": "CASOS · GOLDEN",
                "label": "REQUERIMIENTOS",
                "value": f"{summary['cases']:,}",
                "description": (
                    "Combinaciones tienda–SKU Golden Infaltable con objetivo mayor "
                    "a cero evaluadas por el motor."
                ),
                "tone": "blue",
            },
            {
                "category": "CASOS · GOLDEN",
                "label": "CON ENVÍO",
                "value": f"{summary['served_cases']:,}",
                "description": (
                    "Casos Golden que recibieron al menos una unidad, aunque la "
                    "cobertura haya quedado parcial."
                ),
                "tone": "acid",
            },
            {
                "category": "CASOS · GOLDEN",
                "label": "CUBIERTOS AL 100%",
                "value": f"{summary['full_cases']:,}",
                "description": (
                    "Casos Golden cuya cantidad asignada alcanzó completamente su "
                    "cantidad objetivo."
                ),
                "tone": "acid",
            },
            {
                "category": "CASOS · GOLDEN",
                "label": "SIN ENVÍO",
                "value": f"{summary['not_served_cases']:,}",
                "description": (
                    "Casos Golden con objetivo positivo que no recibieron ninguna "
                    "unidad por alguna regla de corte."
                ),
                "tone": "coral",
            },
            {
                "category": "UNIDADES · GOLDEN",
                "label": "OBJETIVO",
                "value": f"{summary['target_units']:,}",
                "description": (
                    "Suma de las unidades objetivo de todos los casos Golden. No es "
                    "un conteo de tareas."
                ),
            },
            {
                "category": "UNIDADES · GOLDEN",
                "label": "ASIGNADAS",
                "value": f"{summary['assigned_units']:,}",
                "description": (
                    "Unidades de producto realmente asignadas a casos Golden desde "
                    "todos los orígenes."
                ),
                "tone": "acid",
            },
            {
                "category": "COMPLIANCE · CASOS",
                "label": "CASOS COMPLETOS",
                "value": f"{summary['case_compliance_pct']:,.1f}%",
                "description": (
                    "Casos Golden cubiertos al 100% divididos entre todos los casos "
                    "Golden con objetivo positivo."
                ),
                "tone": "blue",
            },
            {
                "category": "COMPLIANCE · UNIDADES",
                "label": "UNIDADES CUBIERTAS",
                "value": f"{summary['unit_compliance_pct']:,.1f}%",
                "description": (
                    "Unidades Golden asignadas divididas entre las unidades Golden "
                    "objetivo. Puede diferir del compliance de casos."
                ),
                "tone": "blue",
            },
        ]
    )

    st.markdown('<span class="section-label">GOLDEN POR CIUDAD</span>', unsafe_allow_html=True)
    compact_city = [
        {
            "CIUDAD": row["CIUDAD"],
            "CASOS": row["CASOS"],
            "COMPLETOS": row["COMPLETOS"],
            "TIENDAS": row["TIENDAS"],
            "OBJETIVO": row["OBJETIVO"],
            "ASIGNADO": row["ASIGNADO"],
            "COMPLIANCE_%": row["COMPLIANCE_%"],
        }
        for row in golden.get("city_rows", [])
    ]
    report_table(
        compact_city,
        column_config={
            "COMPLIANCE_%": st.column_config.NumberColumn(
                "COMPLIANCE %", format="%.1f%%", width="small"
            ),
        },
        max_height=360,
    )

    st.markdown('<span class="section-label">GOLDEN POR ORIGEN</span>', unsafe_allow_html=True)
    report_table(golden.get("origin_rows", []), max_height=300)

    st.markdown('<span class="section-label">GOLDEN POR TIENDA</span>', unsafe_allow_html=True)
    report_table(
        golden.get("store_rows", []),
        column_config={
            "TIENDA": st.column_config.TextColumn("TIENDA", width="large"),
            "COMPLIANCE_%": st.column_config.NumberColumn(
                "COMPLIANCE %", format="%.1f%%", width="small"
            ),
        },
        max_height=520,
    )

    st.markdown(
        '<span class="section-label">DETALLE TIENDA–SKU GOLDEN</span>',
        unsafe_allow_html=True,
    )
    report_table(
        golden.get("detail_rows", []),
        column_config={
            "CIUDAD": st.column_config.TextColumn("CIUDAD", width="small"),
            "TIENDA": st.column_config.TextColumn("TIENDA", width="medium"),
            "SKU": st.column_config.TextColumn("SKU", width="medium"),
            "OBJETIVO": st.column_config.NumberColumn(
                "OBJETIVO", format="%d", width="small"
            ),
            "ASIGNADO": st.column_config.NumberColumn(
                "ASIGNADO", format="%d", width="small"
            ),
            "FALTANTE": st.column_config.NumberColumn(
                "FALTANTE", format="%d", width="small"
            ),
            "ORÍGENES": st.column_config.TextColumn("ORÍGENES", width="small"),
            "BREAKDOWN": st.column_config.TextColumn(
                "BREAKDOWN", width="medium"
            ),
        },
        max_height=650,
    )


def render_planning_analytics(analytics: dict[str, Any]) -> None:
    summary = analytics["summary"]

    st.markdown(
        '<div class="report-title">CIUDADES + TIENDAS.</div>',
        unsafe_allow_html=True,
    )
    st.markdown(
        """
        <div class="report-note">
            PRODUCTOS = SKUs distintos con al menos una unidad asignada · M³ = volumen
            realmente planeado · TAREAS = líneas generadas considerando cada origen.
        </div>
        """,
        unsafe_allow_html=True,
    )

    render_kpi_cards(
        [
            {
                "category": "COBERTURA · GEOGRAFÍA",
                "label": "CIUDADES ATENDIDAS",
                "value": f"{summary['cities_served']:,}",
                "description": (
                    "Número de ciudades distintas con al menos una unidad de producto "
                    "asignada. No incluye insumos."
                ),
            },
            {
                "category": "COBERTURA · TIENDAS",
                "label": "TIENDAS ATENDIDAS",
                "value": f"{summary['stores_served']:,}",
                "description": (
                    "Warehouses destino distintos que reciben al menos una unidad de "
                    "producto normal."
                ),
                "tone": "acid",
            },
            {
                "category": "COBERTURA · SKUs",
                "label": "PRODUCTOS DISTINTOS",
                "value": f"{summary['products_served']:,}",
                "description": (
                    "RETAIL_ID distintos con al menos una unidad asignada en toda la "
                    "planeación. No equivale al número de tareas."
                ),
            },
            {
                "category": "VOLUMEN · PRODUCTO",
                "label": "M³ PLANEADOS",
                "value": f"{summary['m3_assigned']:,.3f}",
                "description": (
                    "Suma del volumen de producto normal asignado: QUANTITY por metros "
                    "cúbicos por unidad. No incluye insumos."
                ),
                "tone": "blue",
            },
        ]
    )

    st.markdown('<span class="section-label">RESUMEN POR CIUDAD</span>', unsafe_allow_html=True)
    report_table(
        analytics["city_rows"],
        column_config={
            "M3": st.column_config.NumberColumn("M³", format="%.3f"),
        },
        max_height=360,
    )

    st.markdown('<span class="section-label">DETALLE POR TIENDA</span>', unsafe_allow_html=True)
    report_table(
        analytics["store_rows"],
        column_config={
            "WAREHOUSE_ID": st.column_config.NumberColumn(
                "WAREHOUSE ID", format="%d"
            ),
            "M3": st.column_config.NumberColumn("M³", format="%.3f"),
        },
        max_height=600,
    )

    st.markdown(
        '<div class="report-title">REPORTE DE REABASTO.</div>',
        unsafe_allow_html=True,
    )
    st.markdown(
        """
        <div class="report-note">
            Un caso representa una combinación tienda–SKU. COMPLIANCE DE CASOS mide
            cuántos casos se cubrieron al 100%; COMPLIANCE DE UNIDADES compara lo
            asignado contra el objetivo final del modelo. El incremento por hardcode se
            presenta separado del ROQ original para no inflar artificialmente el
            cumplimiento. Mantén el cursor un segundo sobre cualquier tarjeta para ver
            exactamente qué mide y cuál es su denominador.
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown('<span class="section-label">CASOS TIENDA–SKU</span>', unsafe_allow_html=True)
    render_kpi_cards(
        [
            {
                "category": "CASOS · BASE",
                "label": "CON RECOMENDACIÓN",
                "value": f"{summary['eligible_cases']:,}",
                "description": (
                    "Combinaciones tienda–SKU con CANTIDAD_OBJETIVO mayor a cero "
                    "después de aplicar ROQ y hardcodes."
                ),
            },
            {
                "category": "CASOS · COMPLETOS",
                "label": "CUBIERTOS AL 100%",
                "value": f"{summary['fully_covered_cases']:,}",
                "description": (
                    "Casos cuya cantidad asignada alcanzó completamente la cantidad "
                    "objetivo. Un caso no es lo mismo que una tarea."
                ),
                "tone": "acid",
            },
            {
                "category": "CASOS · PARCIALES",
                "label": "COBERTURA PARCIAL",
                "value": f"{summary['partial_cases']:,}",
                "description": (
                    "Casos que reciben al menos una unidad, pero no cubren la cantidad "
                    "objetivo completa."
                ),
                "tone": "coral",
            },
            {
                "category": "CASOS · SIN SALIDA",
                "label": "SIN ENVÍO",
                "value": f"{summary['not_assigned_cases']:,}",
                "description": (
                    "Casos con recomendación positiva a los que no se asignó ninguna "
                    "unidad por las reglas de corte."
                ),
                "tone": "coral",
            },
        ]
    )

    st.markdown('<span class="section-label">COMPLIANCE</span>', unsafe_allow_html=True)
    render_kpi_cards(
        [
            {
                "category": "PORCENTAJE · CASOS",
                "label": "COMPLIANCE DE CASOS",
                "value": f"{summary['case_compliance_pct']:,.1f}%",
                "description": (
                    "Casos tienda–SKU cubiertos al 100% dividido entre casos con "
                    "recomendación. Mide cumplimiento completo, no tareas generadas."
                ),
                "tone": "acid",
            },
            {
                "category": "PORCENTAJE · CASOS",
                "label": "ATENCIÓN DE CASOS",
                "value": f"{summary['case_service_pct']:,.1f}%",
                "description": (
                    "Casos tienda–SKU con al menos una unidad asignada dividido entre "
                    "casos con recomendación. Incluye coberturas parciales."
                ),
            },
            {
                "category": "PORCENTAJE · UNIDADES",
                "label": "COMPLIANCE DE UNIDADES",
                "value": f"{summary['unit_compliance_pct']:,.1f}%",
                "description": (
                    "Unidades asignadas dividido entre el objetivo final del modelo, "
                    "incluidos los incrementos de hardcode."
                ),
                "tone": "blue",
            },
            {
                "category": "PORCENTAJE · UNIDADES",
                "label": "COMPLIANCE DEL ROQ ORIGINAL",
                "value": f"{summary['original_roq_compliance_pct']:,.1f}%",
                "description": (
                    "Unidades del ROQ original cubiertas dividido entre las unidades "
                    "del ROQ antes de aplicar mínimos y hardcodes."
                ),
                "tone": "blue",
            },
        ]
    )

    st.markdown('<span class="section-label">UNIDADES</span>', unsafe_allow_html=True)
    render_kpi_cards(
        [
            {
                "category": "UNIDADES · ORIGINALES",
                "label": "ROQ ORIGINAL",
                "value": f"{summary['original_roq_units']:,}",
                "description": (
                    "Suma del ROQ original positivo antes de aplicar el mínimo de tres "
                    "o cualquier hardcode."
                ),
            },
            {
                "category": "UNIDADES · OBJETIVO",
                "label": "OBJETIVO FINAL",
                "value": f"{summary['target_units']:,}",
                "description": (
                    "Unidades solicitadas finalmente por el modelo después de ROQ, "
                    "mínimos y hardcodes."
                ),
                "tone": "acid",
            },
            {
                "category": "UNIDADES · ASIGNADAS",
                "label": "PRODUCTO PLANEADO",
                "value": f"{summary['assigned_units']:,}",
                "description": (
                    "Unidades de producto normal que efectivamente fueron asignadas. "
                    "No incluye insumos."
                ),
                "tone": "blue",
            },
            {
                "category": "UNIDADES · INCREMENTALES",
                "label": "HARDCODE ENVIADO",
                "value": f"{summary['hardcode_assigned_units']:,}",
                "description": (
                    "Unidades realmente asignadas por encima del ROQ original debido "
                    "a las reglas de hardcode."
                ),
                "tone": "coral",
            },
        ]
    )

    st.markdown('<span class="section-label">REGLAS ESPECIALES</span>', unsafe_allow_html=True)
    render_kpi_cards(
        [
            {
                "category": "CASOS · HARDCODE",
                "label": "CASOS CON HARDCODE",
                "value": f"{summary['hardcode_cases']:,}",
                "description": (
                    "Casos donde la cantidad objetivo quedó por encima del ROQ "
                    "original por una regla de negocio."
                ),
            },
            {
                "category": "UNIDADES · HARDCODE",
                "label": "OBJETIVO AÑADIDO",
                "value": f"{summary['hardcode_target_units']:,}",
                "description": (
                    "Diferencia total entre el objetivo final y el ROQ original. "
                    "Representa lo solicitado adicionalmente por hardcodes."
                ),
                "tone": "coral",
            },
            {
                "category": "CASOS · FORECAST 0",
                "label": "FORZADOS A 4",
                "value": f"{summary['forecast_zero_forced_cases']:,}",
                "description": (
                    "Casos con forecast, inventario de apertura y ROQ en cero cuyo "
                    "objetivo fue forzado a cuatro unidades."
                ),
            },
            {
                "category": "CASOS · DÉFICIT",
                "label": "FORZADOS A 3",
                "value": f"{summary['deficit_forced_cases']:,}",
                "description": (
                    "Casos con inventario de apertura menor a la demanda y ROQ cero "
                    "cuyo objetivo fue forzado a tres unidades."
                ),
            },
        ]
    )

    st.markdown('<span class="section-label">STOCKOUTS + PRIORIDADES</span>', unsafe_allow_html=True)
    render_kpi_cards(
        [
            {
                "category": "CASOS · STOCKOUT",
                "label": "STOCKOUTS CON ENVÍO",
                "value": f"{summary['stockout_cases_served']:,}",
                "description": (
                    "Casos tienda–SKU que iniciaron en stockout y reciben al menos "
                    "una unidad."
                ),
                "tone": "acid",
            },
            {
                "category": "SKUs · STOCKOUT",
                "label": "PRODUCTOS DISTINTOS ATENDIDOS",
                "value": f"{summary['stockout_products_served']:,}",
                "description": (
                    "RETAIL_ID distintos en stockout atendidos en al menos una tienda. "
                    "No es un conteo de tareas."
                ),
            },
            {
                "category": "CASOS · STOCKOUT",
                "label": "CUBIERTOS AL 100%",
                "value": f"{summary['stockout_cases_full']:,}",
                "description": (
                    "Casos tienda–SKU en stockout cuya cantidad objetivo fue cubierta "
                    "completamente."
                ),
                "tone": "acid",
            },
            {
                "category": "CASOS · GOLDEN",
                "label": "GOLDEN CON ENVÍO",
                "value": f"{summary['golden_served_cases']:,}",
                "description": (
                    "Casos Golden Infaltables que reciben al menos una unidad, sin "
                    "importar cuántas tareas se generaron."
                ),
                "tone": "blue",
            },
        ]
    )

    render_golden_report(analytics)

    st.markdown('<span class="section-label">REGLAS DE DEMANDA</span>', unsafe_allow_html=True)
    report_table(
        analytics["rule_rows"],
        column_config={
            "COMPLIANCE_UNIDADES_%": st.column_config.NumberColumn(
                "COMPLIANCE UNIDADES %", format="%.1f%%"
            ),
        },
        max_height=360,
    )

    st.markdown('<span class="section-label">STOCKOUTS POR CIUDAD</span>', unsafe_allow_html=True)
    report_table(
        analytics["stockout_city_rows"],
        column_config={
            "ATENCION_%": st.column_config.NumberColumn(
                "ATENCIÓN %", format="%.1f%%"
            ),
        },
        max_height=400,
    )

    st.markdown('<span class="section-label">ASIGNACIÓN POR ORIGEN</span>', unsafe_allow_html=True)
    report_table(
        analytics["source_rows"],
        column_config={
            "WAREHOUSE_SOURCE": st.column_config.NumberColumn(
                "WAREHOUSE SOURCE", format="%d"
            ),
            "M3": st.column_config.NumberColumn("M³", format="%.3f"),
            "UNIDADES_POR_TAREA": st.column_config.NumberColumn(
                "UNIDADES / TAREA", format="%.2f"
            ),
            "PARTICIPACION_TAREAS_%": st.column_config.NumberColumn(
                "% DE TAREAS", format="%.1f%%"
            ),
            "PARTICIPACION_UNIDADES_%": st.column_config.NumberColumn(
                "% DE UNIDADES", format="%.1f%%"
            ),
        },
        max_height=360,
    )

    render_source_analysis(analytics)


def render_results(run: dict[str, Any]) -> None:
    st.markdown('<div class="result-title">PLANEACIÓN LISTA.</div>', unsafe_allow_html=True)
    render_kpi_cards(
        [
            {
                "category": "CASOS · INPUT",
                "label": "REQUERIMIENTOS ÚNICOS RECIBIDOS",
                "value": f"{run.get('input_requirements', run['requirements']):,}",
                "description": (
                    "Combinaciones tienda–SKU únicas leídas del CSV después de "
                    "consolidar duplicados. Incluye tiendas cerradas y ciudades "
                    "bloqueadas antes de aplicar exclusiones."
                ),
                "tone": "acid",
            },
            {
                "category": "CASOS · MODELO",
                "label": "REQUERIMIENTOS EVALUADOS",
                "value": f"{run['requirements']:,}",
                "description": (
                    "Casos tienda–SKU que llegaron al motor después de quitar "
                    "TIENDAS_CERRADAS y los bloqueos opcionales de ciudad."
                ),
            },
            {
                "category": "TAREAS · OPERACIÓN",
                "label": "TAREAS DE ABASTO",
                "value": f"{run['tasks']:,}",
                "description": (
                    "Líneas operativas de transferencia generadas por el modelo. "
                    "Un caso puede crear dos tareas si se divide entre dos orígenes. "
                    "Las líneas de insumos no cuentan aquí."
                ),
                "tone": "blue",
            },
            {
                "category": "UNIDADES · ABASTO",
                "label": "UNIDADES DE PRODUCTO",
                "value": f"{run['units']:,}",
                "description": (
                    "Suma de QUANTITY de las transferencias normales de producto. "
                    "No incluye insumos."
                ),
            },
            {
                "category": "LÍNEAS · INSUMOS",
                "label": "LÍNEAS DE INSUMOS",
                "value": f"{run.get('insumos', {}).get('lines_added', 0):,}",
                "description": (
                    "Filas de INSUMOS anexadas al BulkCD_444 para tiendas que ya "
                    "reciben producto normal desde 444. No consumen tareas."
                ),
            },
            {
                "category": "UNIDADES · INSUMOS",
                "label": "UNIDADES DE INSUMOS",
                "value": f"{run.get('insumos', {}).get('units_added', 0):,}",
                "description": (
                    "Suma de QUANTITY de las líneas de insumos agregadas. Se muestra "
                    "separada para no mezclarla con unidades de producto."
                ),
                "tone": "coral",
            },
        ],
        columns_count=3,
    )

    avl = run.get("avl", {})
    if avl.get("enabled"):
        st.markdown(
            '<span class="section-label">COBERTURA AVL · ÚLTIMA PASADA</span>',
            unsafe_allow_html=True,
        )
        render_kpi_cards(
            [
                {
                    "category": "CASOS · AVL",
                    "label": "STOCKOUTS CON ENVÍO AVL",
                    "value": f"{avl.get('cases_sent', 0):,}",
                    "description": (
                        "Combinaciones tienda–SKU del CATALOGO con stock final "
                        "igual a cero que recibieron producto después de terminar "
                        "la recomendación Fountain9."
                    ),
                    "tone": "acid",
                },
                {
                    "category": "TAREAS · AVL",
                    "label": "TAREAS SOBRANTES UTILIZADAS",
                    "value": f"{avl.get('tasks_added', 0):,}",
                    "description": (
                        "Líneas operativas adicionales utilizadas por AVL. Solo "
                        "consume el remanente entre el máximo configurado y las "
                        "tareas que ya ocupó Fountain9."
                    ),
                    "tone": "blue",
                },
                {
                    "category": "UNIDADES · AVL",
                    "label": "UNIDADES ADICIONALES AVL",
                    "value": f"{avl.get('units_added', 0):,}",
                    "description": (
                        f"Unidades enviadas para cubrir {avl.get('doh', 0):g} DOH "
                        "según ADU. El objetivo mínimo es 3, salvo que el stock "
                        "permitido disponible sea menor."
                    ),
                    "tone": "coral",
                },
            ],
            columns_count=3,
        )
        st.success(
            f"Cobertura AVL a {avl.get('doh', 0):g} DOH: "
            f"{avl.get('cases_sent', 0):,} casos, "
            f"{avl.get('tasks_added', 0):,} tareas y "
            f"{avl.get('units_added', 0):,} unidades adicionales."
        )

    fruver_811 = run.get("fruver_811", {})
    if fruver_811.get("enabled"):
        st.warning(
            "Bloqueo FRUVER 811 aplicado: "
            f"{fruver_811.get('products_with_stock_blocked', 0):,} productos y "
            f"{fruver_811.get('units_blocked', 0):,.0f} unidades de stock del 811 "
            "quedaron fuera de la asignación; los demás orígenes siguieron activos."
        )

    insumos = run.get("insumos", {})
    if insumos.get("lines_added", 0) > 0:
        st.success(
            f"Insumos anexados a BulkCD_444: {insumos['lines_added']:,} líneas, "
            f"{insumos['units_added']:,} unidades y "
            f"{insumos['stores_added']:,} tiendas. No consumen tareas del modelo."
        )

    closed_stores = run.get("closed_stores", {})
    if closed_stores.get("requirements", 0) > 0:
        store_ids = ", ".join(map(str, closed_stores.get("store_ids", [])))
        st.warning(
            "Bloqueo backend TIENDAS_CERRADAS aplicado: se excluyeron "
            f"{closed_stores['requirements']:,} requerimientos de "
            f"{closed_stores['stores']:,} tiendas ({store_ids}) antes de asignar stock."
        )

    city_block = run.get("city_block", {})
    if city_block.get("requirements", 0) > 0:
        city_names = ", ".join(
            item["name"] for item in city_block.get("cities", [])
        )
        st.warning(
            f"Bloqueo de ciudad aplicado: {city_names}. Se excluyeron "
            f"{city_block['requirements']:,} requerimientos, "
            f"{city_block['stores']:,} tiendas y "
            f"{city_block['products']:,} productos antes de asignar stock."
        )

    st.markdown('<span class="section-label">DESCARGAR TODO</span>', unsafe_allow_html=True)
    zip_path = Path(run["zip"])
    st.download_button(
        "Descargar planeación completa (.zip)",
        data=zip_path.read_bytes(),
        file_name=zip_path.name,
        mime="application/zip",
        use_container_width=True,
    )

    st.markdown('<span class="section-label">ARCHIVOS INDIVIDUALES</span>', unsafe_allow_html=True)
    columns = st.columns(min(max(len(run["files"]), 1), 3))
    for index, raw_path in enumerate(run["files"]):
        path = Path(raw_path)
        mime = (
            "text/csv"
            if path.suffix.lower() == ".csv"
            else "application/pdf"
            if path.suffix.lower() == ".pdf"
            else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        with columns[index % len(columns)]:
            st.markdown(f'<span class="file-pill">{path.name}</span>', unsafe_allow_html=True)
            st.download_button(
                f"Descargar {path.name}",
                data=path.read_bytes(),
                file_name=path.name,
                mime=mime,
                key=f"download_{index}_{path.name}",
                use_container_width=True,
            )

    st.markdown('<span class="section-label">BREAKDOWN</span>', unsafe_allow_html=True)
    breakdown = ordered_breakdown_rows(run["status_counts"])
    st.dataframe(breakdown, use_container_width=True, hide_index=True)

    if run.get("analytics"):
        render_planning_analytics(run["analytics"])

    if run["warnings"]:
        with st.expander(f"Advertencias de calidad ({len(run['warnings'])})"):
            for warning in run["warnings"]:
                st.warning(warning)
    with st.expander("Log técnico de la ejecución"):
        st.code(run["logs"] or "Ejecución completada sin mensajes adicionales.")


def main() -> None:
    inject_styles()

    st.markdown(
        """
        <section class="hero">
            <span class="hero-kicker">ABASTO / MX / WEB</span>
            <h1>TRANSFER<br>PLANNER.</h1>
            <p>Sube el requerimiento diario, define los orígenes y ejecuta el mismo motor sin editar código. La base se consulta y valida automáticamente.</p>
        </section>
        """,
        unsafe_allow_html=True,
    )

    st.markdown('<span class="section-label">01 — BASE DE DATOS</span>', unsafe_allow_html=True)
    database_bytes: bytes | None = None
    database_health: dict[str, Any] | None = None
    city_labels: dict[str, str] = {}
    try:
        with st.spinner("Validando fuentes de información…"):
            database_bytes = fetch_public_database()
            database_health = inspect_database(database_bytes)
            city_labels = extract_available_cities(database_bytes)
        render_database_health(database_health)
    except Exception as exc:
        st.markdown(
            """
            <div class="database-status review">
                <div class="database-title">BASE DE DATOS — OFFLINE</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.error(f"No fue posible validar la base de datos: {exc}")

    if st.button("VOLVER A VALIDAR LA BASE", use_container_width=False):
        fetch_public_database.clear()
        st.rerun()

    st.markdown('<span class="section-label">02 — ARCHIVO DE PLANEACIÓN</span>', unsafe_allow_html=True)
    uploaded_plan = st.file_uploader(
        "Plan diario (.csv)",
        type=["csv"],
        max_upload_size=MAX_UPLOAD_MB,
        help="El archivo debe conservar las columnas originales del plan.",
    )
    if uploaded_plan is not None:
        st.success(
            f"{uploaded_plan.name} · {uploaded_plan.size / (1024 ** 2):,.1f} MB"
        )

    st.markdown('<span class="section-label">03 — VARIABLES</span>', unsafe_allow_html=True)
    run_date = datetime.now(ZoneInfo("America/Mexico_City")).date()
    default_origins = [
        origin
        for origin in engine.CONFIG.origin_warehouses
        if origin in ORIGIN_WAREHOUSES
    ] or [811, 834]

    with st.form("planning_config"):
        left, right = st.columns([2, 1])
        with left:
            selected_origins = st.multiselect(
                "Warehouses origen — selecciona en orden de prioridad",
                options=list(ORIGIN_WAREHOUSES),
                default=default_origins,
                format_func=format_origin,
                help=(
                    "El motor consumirá stock en el orden seleccionado. Para cambiar "
                    "la prioridad, elimina las opciones y vuelve a elegirlas."
                ),
            )
        with right:
            max_tasks = st.number_input(
                "Máximo de tareas",
                min_value=0,
                max_value=1_000_000,
                value=int(engine.CONFIG.max_tasks),
                step=500,
            )

        selected_blocked_cities = st.multiselect(
            "Bloquear ciudades completas — opcional",
            options=list(city_labels),
            default=[],
            format_func=lambda city: city_labels.get(city, city),
            help=(
                "No se generarán envíos hacia las ciudades seleccionadas. Sus "
                "requerimientos se eliminan antes de asignar, por lo que ese stock "
                "queda disponible para otras tiendas."
            ),
        )
        if selected_blocked_cities:
            selected_names = ", ".join(
                city_labels.get(city, city) for city in selected_blocked_cities
            )
            st.warning(f"Se bloqueará completamente: {selected_names}")

        include_insumos = st.toggle(
            "Agregar insumos al BulkCD_444",
            value=True,
            help=(
                "Cuando está activo, anexa los insumos de Aleph únicamente a las "
                "tiendas que ya reciben producto normal desde el warehouse 444. "
                "Los insumos no consumen tareas, stock ni capacidad."
            ),
        )

        block_fruver_811 = st.toggle(
            "Bloquear envíos FRUVER desde el warehouse 811",
            value=False,
            help=(
                "Cuando está activo y el 811 es un origen seleccionado, su stock "
                "de productos clasificados como FRUVER queda fuera de la asignación. "
                "El motor puede cubrirlos desde los demás orígenes disponibles."
            ),
        )

        avl_left, avl_right = st.columns([2, 1])
        with avl_left:
            include_avl_fill = st.toggle(
                "Completar tareas disponibles con cobertura AVL",
                value=False,
                help=(
                    "Después de cubrir Fountain9, busca productos del catálogo con stock "
                    "final igual a cero en la tienda y utiliza únicamente las "
                    "tareas sobrantes. Mantiene capacidad, rutas, bloqueos, "
                    "rackeados y prioridad de orígenes."
                ),
            )
        with avl_right:
            avl_doh = st.number_input(
                "DOH para cobertura AVL",
                min_value=0.5,
                max_value=30.0,
                value=3.0,
                step=0.5,
                disabled=not include_avl_fill,
                help=(
                    "Cantidad objetivo = ADU × DOH, redondeada hacia arriba. "
                    "Se envían al menos 3 unidades; si el stock permitido no "
                    "alcanza, se manda lo disponible."
                ),
            )

        submitted = st.form_submit_button(
            "EJECUTAR PLANEACIÓN →", use_container_width=True
        )

    if submitted:
        if uploaded_plan is None:
            st.error("Primero sube el CSV del plan diario.")
            st.stop()
        if not selected_origins:
            st.error("Selecciona al menos un warehouse origen.")
            st.stop()
        try:
            origins = tuple(selected_origins)
            with st.status("Ejecutando motor de planeación…", expanded=True) as status:
                st.write("Guardando el CSV cargado de forma temporal…")
                st.write("Validando la versión actual de la base de datos…")
                fetch_public_database.clear()
                database_bytes = fetch_public_database()
                database_health = inspect_database(database_bytes)
                if not database_health["online"]:
                    raise RuntimeError(
                        "La base de datos tiene una o más fuentes con error. "
                        "Abre el panel de estado y corrige las hojas indicadas."
                    )
                if selected_blocked_cities:
                    blocked_names = ", ".join(
                        city_labels.get(city, city)
                        for city in selected_blocked_cities
                    )
                    st.write(f"Excluyendo ciudades bloqueadas: {blocked_names}…")
                if include_insumos and 444 in origins:
                    st.write("Preparando insumos para las rutas activas del 444…")
                elif include_insumos:
                    st.write(
                        "Insumos activos, pero sin efecto porque el origen 444 no "
                        "está seleccionado."
                    )
                else:
                    st.write("Envío de insumos desactivado para esta corrida.")
                if block_fruver_811 and 811 in origins:
                    st.write(
                        "Bloqueando el stock FRUVER del origen 811 y habilitando "
                        "la reasignación desde otros orígenes…"
                    )
                elif block_fruver_811:
                    st.write(
                        "Bloqueo FRUVER 811 activo, pero sin efecto porque el 811 "
                        "no está seleccionado como origen."
                    )
                st.write("Calculando demanda, stock, capacidad y tareas…")
                if include_avl_fill:
                    st.write(
                        f"Reservando la última pasada para cobertura AVL a "
                        f"{avl_doh:g} DOH…"
                    )
                run = execute_planning(
                    uploaded_plan=uploaded_plan,
                    database_bytes=database_bytes,
                    origins=origins,
                    max_tasks=int(max_tasks),
                    run_date=run_date,
                    blocked_cities=tuple(selected_blocked_cities),
                    include_insumos=include_insumos,
                    include_avl_fill=include_avl_fill,
                    avl_doh=float(avl_doh),
                    block_fruver_811=block_fruver_811,
                )
                status.update(label="Planeación finalizada", state="complete", expanded=False)
            st.session_state["last_run"] = run
        except Exception as exc:
            st.session_state.pop("last_run", None)
            st.error(f"No se pudo completar la planeación: {exc}")
            st.info(
                "Revisa el CSV, el panel de estado y que los warehouses origen existan en TIENDA."
            )

    last_run = st.session_state.get("last_run")
    if last_run and Path(last_run["zip"]).exists():
        render_results(last_run)


if __name__ == "__main__":
    main()
